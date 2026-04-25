"""Smoke test for the professional audit Excel generator.

We don't try to assert pixel-perfect formatting — that would brittle the
test against xlsxwriter version bumps. Instead we verify:

* All three expected sheets exist with the right names
* The Resumen sheet shows the right KPIs in the right cells
* The Detalle sheet has one row per audit + a header row
* The Banderas Rojas sheet only lists rows where ``needs_review`` is True
"""
from __future__ import annotations

from openpyxl import load_workbook

from secop_ii.audit import (
    VERDICT_AGREE_NO,
    VERDICT_NOTE_ONLY,
    RowAudit,
)
from secop_ii.excel_pro import build_audit_workbook


def _audit(fila: int, *, verdict: str, needs_review: bool, **kw) -> RowAudit:
    defaults = dict(
        process_id=f"CO1.PPI.{fila:08d}",
        objeto="objeto demo",
        note_text="nota demo",
        note_says_modif=False,
        note_says_no_leg=False,
        api_says_modif=False,
        api_count=0,
        docs_says_modif=False,
        docs_mod_count=0,
        docs_leg_count=0,
        docs_mod_list="",
        docs_leg_list="",
    )
    defaults.update(kw)
    return RowAudit(fila=fila, verdict=verdict, needs_review=needs_review, **defaults)


class TestBuildAuditWorkbook:
    def test_three_sheets_with_expected_names(self, tmp_path):
        audits = [
            _audit(2, verdict=VERDICT_AGREE_NO, needs_review=False),
            _audit(13, verdict=VERDICT_NOTE_ONLY, needs_review=True,
                   note_text="Modificatorio 1 del 6/03/2024",
                   note_says_modif=True),
        ]
        out = tmp_path / "out.xlsx"
        info = build_audit_workbook(audits, out, excel_source="demo.xlsx")

        assert out.is_file()
        assert info.total_rows == 2
        assert info.needs_review == 1

        wb = load_workbook(out, data_only=True)
        assert wb.sheetnames == ["Resumen", "Detalle", "Banderas Rojas"]

    def test_resumen_kpis(self, tmp_path):
        audits = [
            _audit(1, verdict=VERDICT_AGREE_NO, needs_review=False),
            _audit(2, verdict=VERDICT_AGREE_NO, needs_review=False),
            _audit(3, verdict=VERDICT_NOTE_ONLY, needs_review=True),
        ]
        out = tmp_path / "out.xlsx"
        build_audit_workbook(audits, out, excel_source="x.xlsx")
        ws = load_workbook(out, data_only=True)["Resumen"]
        assert ws["A4"].value == "Archivo origen"
        assert ws["B4"].value == "x.xlsx"
        assert ws["B6"].value == 3
        assert ws["B7"].value == 1

    def test_detalle_has_one_row_per_audit_plus_header(self, tmp_path):
        audits = [_audit(i, verdict=VERDICT_AGREE_NO, needs_review=False) for i in range(2, 12)]
        out = tmp_path / "out.xlsx"
        build_audit_workbook(audits, out)
        ws = load_workbook(out, data_only=True)["Detalle"]
        # 1 header + 10 audits
        assert ws.max_row == 11
        assert ws.cell(row=1, column=1).value == "Fila"
        assert ws.cell(row=2, column=1).value == 2
        # Autofilter must span the full data range
        assert ws.auto_filter.ref is not None

    def test_banderas_rojas_only_lists_needs_review(self, tmp_path):
        audits = [
            _audit(1, verdict=VERDICT_AGREE_NO, needs_review=False),
            _audit(2, verdict=VERDICT_NOTE_ONLY, needs_review=True),
            _audit(3, verdict=VERDICT_NOTE_ONLY, needs_review=True),
            _audit(4, verdict=VERDICT_AGREE_NO, needs_review=False),
        ]
        out = tmp_path / "out.xlsx"
        build_audit_workbook(audits, out)
        ws = load_workbook(out, data_only=True)["Banderas Rojas"]
        # Header on row 4, then 2 flagged rows on 5-6
        assert ws.cell(row=5, column=1).value == 2
        assert ws.cell(row=6, column=1).value == 3
        # Row 1 (the one that wasn't flagged) must NOT appear in this sheet
        all_filas = [
            ws.cell(row=r, column=1).value for r in range(5, 7)
        ]
        assert 1 not in all_filas
        assert 4 not in all_filas
