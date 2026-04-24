"""End-to-end test for the Excel update pipeline with a fake Socrata client."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook as _load

from secop_ii import orchestrator as orch_module
from secop_ii.excel_io import backup_workbook, detect_url_column, read_headers
from secop_ii.orchestrator import process_workbook

SAMPLE_URLS = [
    "https://community.secop.gov.co/Public/Tendering/ContractNoticePhases/View"
    "?PPI=CO1.PPI.46305103&isFromPublicArea=True",
    "https://community.secop.gov.co/Public/Tendering/OpportunityDetail/Index"
    "?noticeUID=CO1.NTC.9999001",
    "",  # blank row — should be skipped
    "esto no es una url",  # invalid
]


def _make_sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Procesos"
    ws.append(["#", "Entidad", "URL del proceso", "Objeto"])
    ws.append([1, "FEAB", SAMPLE_URLS[0], "FEAB-EB-0001-2026"])
    ws.append([2, "Alcaldía Ejemplo", SAMPLE_URLS[1], "Suministro papelería"])
    ws.append([3, "Otra", SAMPLE_URLS[2], "Sin URL"])
    ws.append([4, "Tercera", SAMPLE_URLS[3], "URL rota"])
    wb.save(path)


class _FakeClient:
    """Minimal drop-in for SecopClient used in pipeline tests."""

    def __init__(self, *_, **__):
        self._proc = {
            "CO1.PPI.46305103": {
                "id_del_proceso": "CO1.PPI.46305103",
                "adendas": "",
            },
            "CO1.NTC.9999001": {
                "id_del_proceso": "CO1.NTC.9999001",
                "adendas": "Adenda 1; Adenda 2",
            },
        }
        self._contracts = {
            "CO1.NTC.9999001": [
                {
                    "id_contrato": "CO1.PCCNTR.5551234",
                    "valor_pagado_adiciones": "8000000",
                    "dias_adicionados": "30",
                }
            ],
        }
        self._adiciones = {
            "CO1.PCCNTR.5551234": [
                {
                    "id_contrato": "CO1.PCCNTR.5551234",
                    "tipo_modificacion": "Adición",
                    "descripci_n": "Adición por mayor cantidad",
                    "fecha_registro": "2026-02-10T00:00:00.000",
                    "valor_adicion": "8000000",
                },
                {
                    "id_contrato": "CO1.PCCNTR.5551234",
                    "tipo_modificacion": "Prórroga",
                    "descripci_n": "Prórroga de 30 días",
                    "fecha_registro": "2026-03-05T00:00:00.000",
                    "valor_adicion": "0",
                },
            ],
        }

    def get_proceso(self, process_id, url=None):
        return self._proc.get(process_id)

    def get_contratos(self, process_id):
        return self._contracts.get(process_id, [])

    def get_adiciones(self, id_contrato):
        return self._adiciones.get(id_contrato, [])

    def build_query_url(self, dataset_id, *, where=None, limit=10):
        return (
            f"https://www.datos.gov.co/resource/{dataset_id}.json?"
            f"$limit={limit}&$where={where or ''}"
        )


@pytest.fixture
def sample_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "procesos.xlsx"
    _make_sample_workbook(path)
    return path


class TestBackupAndDetect:
    def test_backup_creates_timestamped_copy(self, sample_workbook: Path):
        backup = backup_workbook(sample_workbook)
        assert backup.exists()
        assert backup != sample_workbook
        assert backup.name.startswith("procesos.backup_")
        assert backup.suffix == ".xlsx"

    def test_detect_url_column_by_header_name(self, sample_workbook: Path):
        wb = _load(sample_workbook)
        idx = detect_url_column(wb.active)
        # "URL del proceso" is the third column
        assert idx == 3


class TestProcessWorkbook:
    def test_full_pipeline_updates_rows_and_adds_columns(
        self, sample_workbook: Path
    ):
        with patch.object(orch_module, "SecopClient", _FakeClient):
            report = process_workbook(
                sample_workbook,
                app_token=None,
                do_backup=True,
            )

        assert report.total == 3  # 3 URLs, the blank row is skipped
        assert report.errors == 1  # the "esto no es una url" row
        assert report.ok == 2
        assert report.with_modificatorio == 1
        assert report.without_modificatorio == 1

        wb = _load(sample_workbook)
        headers = read_headers(wb.active)
        # Modificatorios columns
        assert "¿Hubo modificatorio?" in headers
        assert "# modificatorios" in headers
        # Orchestrator columns
        assert "Estado actualización" in headers
        assert "Última actualización" in headers
        # Auditoría / verificación columns
        assert "ID identificado" in headers
        assert "Fase en SECOP" in headers
        assert "Entidad en SECOP" in headers
        assert "Link verificación API" in headers

        # Row 2 -> PPI.46305103 -> "No" modificatorios
        assert _cell(wb.active, 2, headers["¿Hubo modificatorio?"]) == "No"
        # Row 3 -> NTC.9999001 -> "Sí" modificatorios
        assert _cell(wb.active, 3, headers["¿Hubo modificatorio?"]) == "Sí"
        # Row 5 -> invalid URL -> status url_invalida
        assert _cell(wb.active, 5, headers["Estado actualización"]) == "url_invalida"
        # Audit link must be a datos.gov.co URL on OK rows
        link_cell = _cell(wb.active, 2, headers["Link verificación API"])
        assert link_cell and "datos.gov.co" in link_cell

    def test_backup_is_created_next_to_original(self, sample_workbook: Path):
        with patch.object(orch_module, "SecopClient", _FakeClient):
            report = process_workbook(sample_workbook)
        assert report.backup_path is not None
        assert report.backup_path.exists()
        assert report.backup_path.parent == sample_workbook.parent

    def test_dry_run_does_not_modify_file(self, sample_workbook: Path):
        original_mtime = sample_workbook.stat().st_mtime
        with patch.object(orch_module, "SecopClient", _FakeClient):
            process_workbook(sample_workbook, dry_run=True, do_backup=False)
        assert sample_workbook.stat().st_mtime == original_mtime


def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value
