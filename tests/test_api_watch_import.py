"""Tests for ``POST /watch/import-from-excel``.

We don't hit the real SECOP — the endpoint calls
``_client.resolve_notice_uid`` for each parsed URL, so we monkeypatch it
to a no-op. The watch list lives at ``.cache/watched_urls.json``; we
redirect it to a tmp path so tests don't pollute real state.

Verifies:

* Auto-detects header row (1 vs 4 — the legacy ``FEAB 2018-2021`` layout)
* Auto-detects the LINK column case-insensitively
* Skips URLs that aren't SECOP
* Dedups within the run AND against an already-populated watch list
* Per-sheet stats add up to the global counts
"""
from __future__ import annotations

from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook


@pytest.fixture
def fixture_workbook(tmp_path: Path) -> Path:
    """Build a 3-sheet workbook that exercises every layout we expect."""
    wb = Workbook()
    # Sheet 1: row-1 headers, two valid SECOP URLs (one of which is dup
    # of sheet 2), one empty, one non-SECOP URL.
    s1 = wb.active
    s1.title = "FEAB 2026"
    s1.append(["#", "OBJETO", "LINK"])
    s1.append([1, "Contrato A",
              "https://community.secop.gov.co/Public/Tendering/"
              "OpportunityDetail/Index?noticeUID=CO1.NTC.1111111"])
    s1.append([2, "Contrato B",
              "https://community.secop.gov.co/Public/Tendering/"
              "OpportunityDetail/Index?noticeUID=CO1.NTC.2222222"])
    s1.append([3, "Contrato vacio", ""])
    s1.append([4, "URL externa", "https://www.google.com/search?q=foo"])

    # Sheet 2: row-1 headers — one URL is a dup of sheet 1 (NTC.1111111),
    # one new.
    s2 = wb.create_sheet("FEAB 2025")
    s2.append(["#", "OBJETO", "LINK"])
    s2.append([1, "Contrato A (otra hoja)",
              "https://community.secop.gov.co/Public/Tendering/"
              "OpportunityDetail/Index?noticeUID=CO1.NTC.1111111"])
    s2.append([2, "Contrato C",
              "https://community.secop.gov.co/Public/Tendering/"
              "OpportunityDetail/Index?noticeUID=CO1.NTC.3333333"])

    # Sheet 3: row-4 headers (legacy FEAB 2018-2021 layout). LINK is in
    # an arbitrary column to verify we find it by name, not by position.
    s3 = wb.create_sheet("FEAB 2018-2021")
    s3.append(["PROCESO GESTIÓN CONTRACTUAL"])
    s3.append(["FORMATO"])
    s3.append([])
    s3.append(["#", "OBJETO", "LINK", "EXTRA"])
    s3.append([1, "Contrato D",
              "https://community.secop.gov.co/Public/Tendering/"
              "OpportunityDetail/Index?noticeUID=CO1.NTC.4444444",
              "x"])

    out = tmp_path / "fixture.xlsx"
    wb.save(out)
    return out


@pytest.fixture
def client(monkeypatch, tmp_path: Path) -> TestClient:
    """Build a TestClient with the watch list and SECOP client patched
    so the test runs offline and doesn't pollute real cache files."""
    from secop_ii import api as api_module

    # Redirect watch list to tmp
    monkeypatch.setattr(api_module, "_WATCH_PATH",
                        tmp_path / "watched_urls.json")

    # Block any SECOP network call. resolve_notice_uid is the only one
    # the importer touches.
    monkeypatch.setattr(api_module._client, "resolve_notice_uid",
                        lambda pid, url=None: None)

    return TestClient(api_module.app)


class TestWatchImportFromExcel:
    def test_imports_all_unique_secop_urls(self, client, fixture_workbook):
        res = client.post("/watch/import-from-excel",
                         json={"workbook": str(fixture_workbook)})
        assert res.status_code == 200
        body = res.json()

        # 4 unique SECOP processes across the 3 sheets:
        #   NTC.1111111 (sheets 1+2 → 1 unique item, 2 appearances)
        #   NTC.2222222 (sheet 1)
        #   NTC.3333333 (sheet 2)
        #   NTC.4444444 (sheet 3 — row-4 headers)
        assert body["added_new"] == 4
        assert body["merged"] == 1  # the cross-sheet 1111111 second hit
        assert body["total_unique"] == 4
        # Total appearances = sum of every (sheet, row) hit = 5
        assert body["total_appearances"] == 5

    def test_per_sheet_counts_match_excel(self, client, fixture_workbook):
        """The CARDINAL test: per-sheet counts must equal what the Dra
        sees if she opens that sheet in Excel."""
        body = client.post("/watch/import-from-excel",
                          json={"workbook": str(fixture_workbook)}).json()
        per_sheet = body["per_sheet"]

        # Sheet "FEAB 2026" has 2 SECOP URLs (1111111 + 2222222), both
        # new the first time. found=2 mirrors what the Dra sees.
        assert per_sheet["FEAB 2026"]["found"] == 2
        assert per_sheet["FEAB 2026"]["added_new"] == 2

        # Sheet "FEAB 2025" has 2 SECOP URLs (1111111 dup-of-2026 +
        # 3333333 new). Found=2, but added_new=1 + merged=1.
        assert per_sheet["FEAB 2025"]["found"] == 2
        assert per_sheet["FEAB 2025"]["added_new"] == 1
        assert per_sheet["FEAB 2025"]["merged"] == 1

        assert per_sheet["FEAB 2018-2021"]["found"] == 1
        assert per_sheet["FEAB 2018-2021"]["added_new"] == 1

    def test_filter_by_sheet_returns_excel_count(
            self, client, fixture_workbook):
        """If the Dra filters by 'FEAB 2026', she should see exactly the
        2 processes that are in that sheet — including the cross-sheet
        one (1111111) that ALSO appears on FEAB 2025."""
        client.post("/watch/import-from-excel",
                   json={"workbook": str(fixture_workbook)})
        items = client.get("/watch").json()["items"]
        on_2026 = [it for it in items if "FEAB 2026" in it.get("sheets", [])]
        on_2025 = [it for it in items if "FEAB 2025" in it.get("sheets", [])]
        on_2018 = [it for it in items
                  if "FEAB 2018-2021" in it.get("sheets", [])]
        # These match the Excel exactly: 2, 2, 1.
        assert len(on_2026) == 2
        assert len(on_2025) == 2
        assert len(on_2018) == 1

    def test_running_twice_is_idempotent(
            self, client, fixture_workbook):
        first = client.post("/watch/import-from-excel",
                           json={"workbook": str(fixture_workbook)}).json()
        assert first["added_new"] == 4
        assert first["total_appearances"] == 5

        second = client.post("/watch/import-from-excel",
                            json={"workbook": str(fixture_workbook)}).json()
        # No new items, no new appearances — every (sheet, row, url)
        # already on file gets counted as `already_recorded`.
        assert second["added_new"] == 0
        assert second["merged"] == 0
        assert second["already_recorded"] == 5
        assert second["total_unique"] == 4
        assert second["total_appearances"] == 5

    def test_missing_workbook_returns_404(self, client, tmp_path):
        res = client.post("/watch/import-from-excel",
                         json={"workbook": str(tmp_path / "nope.xlsx")})
        assert res.status_code == 404

    def test_default_workbook_used_when_omitted(
            self, client, monkeypatch, tmp_path, fixture_workbook):
        from secop_ii import api as api_module
        monkeypatch.setattr(api_module, "_DEFAULT_WORKBOOK",
                           str(fixture_workbook))
        res = client.post("/watch/import-from-excel", json={})
        assert res.status_code == 200
        assert res.json()["added_new"] == 4

    def test_persists_to_watched_urls_json(
            self, client, fixture_workbook, monkeypatch, tmp_path):
        client.post("/watch/import-from-excel",
                   json={"workbook": str(fixture_workbook)})
        items = client.get("/watch").json()["items"]
        assert len(items) == 4
        pids = {it["process_id"] for it in items}
        assert pids == {
            "CO1.NTC.1111111", "CO1.NTC.2222222",
            "CO1.NTC.3333333", "CO1.NTC.4444444",
        }
        # Each note should record which sheet it came from
        sheet_notes = {it["note"] for it in items}
        assert any("FEAB 2026" in n for n in sheet_notes)
        assert any("FEAB 2018-2021" in n for n in sheet_notes)


# ---------------------------------------------------------------------------
# Consecutivo FEAB extraction (CARDINAL · 2026-04-28)
#
# La Dra habla por consecutivo (CONTRATO-FEAB-0001-2024), no por process_id.
# El importer debe extraerlo del Excel para que la app hable SU idioma.
# Modelo 1↔N: una URL puede tener hasta 13 contratos (subasta con N
# adjudicaciones) → persistir como lista, no scalar.
# ---------------------------------------------------------------------------


class TestConsecutivoExtraction:
    """Tests directos del helper `_extract_consecutivo_feab`."""

    def test_literal_format_normalizes_padding(self):
        from secop_ii.api import _extract_consecutivo_feab
        assert _extract_consecutivo_feab(
            "CONTRATO-FEAB-0001-2024", None
        ) == "CONTRATO-FEAB-0001-2024"
        # Sin padding del Excel → padded a 4 dígitos
        assert _extract_consecutivo_feab(
            "CONTRATO-FEAB-1-2024", None
        ) == "CONTRATO-FEAB-0001-2024"
        # Con espacios en lugar de guiones
        assert _extract_consecutivo_feab(
            "CONTRATO FEAB 42 2023", None
        ) == "CONTRATO-FEAB-0042-2023"

    def test_numeric_format_uses_fallback_vigencia(self):
        """FEAB 2018-2021 trae solo un entero en col 2 + vigencia col 3."""
        from secop_ii.api import _extract_consecutivo_feab
        assert _extract_consecutivo_feab(1, "2018") == "CONTRATO-FEAB-0001-2018"
        assert _extract_consecutivo_feab("42", "2020") == "CONTRATO-FEAB-0042-2020"
        # Float (Excel siempre devuelve float para int cells)
        assert _extract_consecutivo_feab(7.0, "2019") == "CONTRATO-FEAB-0007-2019"

    def test_returns_none_for_invalid_input(self):
        from secop_ii.api import _extract_consecutivo_feab
        assert _extract_consecutivo_feab(None, "2024") is None
        assert _extract_consecutivo_feab("", "2024") is None
        assert _extract_consecutivo_feab("hola mundo", "2024") is None
        # Numeric mode requiere fallback_vigencia
        assert _extract_consecutivo_feab(5, None) is None
        assert _extract_consecutivo_feab(5, "no-es-año") is None
        # Out of plausible range
        assert _extract_consecutivo_feab(0, "2018") is None


@pytest.fixture
def fixture_workbook_with_consecutivo(tmp_path: Path) -> Path:
    """Workbook con consecutivos en col 1 (formato literal) y formato
    numérico (col 2 con vigencia col 3) para FEAB 2018-2021."""
    wb = Workbook()
    # Sheet con header row 1 + col 1 = numero contrato literal
    s1 = wb.active
    s1.title = "FEAB 2024"
    s1.append(["2. NÚMERO DE CONTRATO", "3.VIGENCIA", "OBJETO", "LINK"])
    s1.append(["CONTRATO-FEAB-0001-2024", 2024, "Contrato A",
              "https://community.secop.gov.co/Public/Tendering/"
              "OpportunityDetail/Index?noticeUID=CO1.NTC.1111111"])
    # Mismo URL — modelo 1↔N · debe acumular consecutivos en lista
    s1.append(["CONTRATO-FEAB-0002-2024", 2024, "Contrato B (mismo proceso)",
              "https://community.secop.gov.co/Public/Tendering/"
              "OpportunityDetail/Index?noticeUID=CO1.NTC.1111111"])
    # URL distinta sin consecutivo (proceso sin contrato firmado)
    s1.append([None, 2024, "Sin contrato",
              "https://community.secop.gov.co/Public/Tendering/"
              "OpportunityDetail/Index?noticeUID=CO1.NTC.2222222"])

    # Sheet con header row 4 + col 2 numérica (FEAB 2018-2021 layout)
    s2 = wb.create_sheet("FEAB 2018-2021")
    s2.append(["PROCESO GESTIÓN CONTRACTUAL"])
    s2.append(["FORMATO"])
    s2.append([])
    s2.append(["#", "2. NÚMERO DE CONTRATO", "3.VIGENCIA", "OBJETO", "LINK"])
    s2.append([1, 1, 2018, "Contrato D",
              "https://community.secop.gov.co/Public/Tendering/"
              "OpportunityDetail/Index?noticeUID=CO1.NTC.4444444"])

    out = tmp_path / "fixture_consec.xlsx"
    wb.save(out)
    return out


class TestImportExtractsConsecutivo:
    def test_consecutivo_persisted_as_list(
            self, client, fixture_workbook_with_consecutivo):
        """Cada URL termina con un `numero_contrato_excel: list[str]`."""
        client.post("/watch/import-from-excel",
                   json={"workbook": str(fixture_workbook_with_consecutivo)})
        items = client.get("/watch").json()["items"]
        by_pid = {it["process_id"]: it for it in items}

        # NTC.1111111 tuvo 2 consecutivos asociados (modelo 1↔N)
        assert by_pid["CO1.NTC.1111111"]["numero_contrato_excel"] == [
            "CONTRATO-FEAB-0001-2024",
            "CONTRATO-FEAB-0002-2024",
        ]
        # NTC.2222222 sin consecutivo en Excel → lista vacía honesta
        assert by_pid["CO1.NTC.2222222"]["numero_contrato_excel"] == []
        # NTC.4444444 con consecutivo numérico construido virtualmente
        assert by_pid["CO1.NTC.4444444"]["numero_contrato_excel"] == [
            "CONTRATO-FEAB-0001-2018",
        ]

    def test_per_sheet_consec_extracted_count(
            self, client, fixture_workbook_with_consecutivo):
        body = client.post(
            "/watch/import-from-excel",
            json={"workbook": str(fixture_workbook_with_consecutivo)},
        ).json()
        per_sheet = body["per_sheet"]
        # FEAB 2024: 3 URLs encontradas (incluyendo dup), 2 con consecutivo
        assert per_sheet["FEAB 2024"]["found"] == 3
        assert per_sheet["FEAB 2024"]["consec_extracted"] == 2
        # FEAB 2018-2021: 1 URL, 1 consecutivo construido
        assert per_sheet["FEAB 2018-2021"]["consec_extracted"] == 1

    def test_idempotent_does_not_duplicate_consecutivos(
            self, client, fixture_workbook_with_consecutivo):
        """Re-correr el import NO duplica consecutivos en la lista."""
        for _ in range(2):
            client.post("/watch/import-from-excel",
                       json={"workbook": str(fixture_workbook_with_consecutivo)})
        items = client.get("/watch").json()["items"]
        by_pid = {it["process_id"]: it for it in items}
        # Sigue siendo 2 consecutivos, no 4
        assert by_pid["CO1.NTC.1111111"]["numero_contrato_excel"] == [
            "CONTRATO-FEAB-0001-2024",
            "CONTRATO-FEAB-0002-2024",
        ]

    def test_backfill_consecutivo_for_preexisting_items(
            self, client, monkeypatch, tmp_path,
            fixture_workbook_with_consecutivo):
        """REGRESSION (2026-04-28): el watch list ya existe con appearances
        pero sin numero_contrato_excel (legacy). Re-importar el Excel debe
        backfillear los consecutivos sin duplicar appearances."""
        from secop_ii import api as api_module
        import json

        # Simular el watch list legacy: 1 item ya registrado con la URL
        # y appearance de FEAB 2024 row 2 — pero SIN numero_contrato_excel.
        legacy_item = {
            "url": "https://community.secop.gov.co/Public/Tendering/"
                   "OpportunityDetail/Index?noticeUID=CO1.NTC.1111111",
            "process_id": "CO1.NTC.1111111",
            "notice_uid": None,
            "sheets": ["FEAB 2024"],
            "vigencias": ["2024"],
            "appearances": [{
                "sheet": "FEAB 2024",
                "vigencia": "2024",
                "row": 2,
                "url": "https://community.secop.gov.co/Public/Tendering/"
                       "OpportunityDetail/Index?noticeUID=CO1.NTC.1111111",
            }],
            "added_at": "2026-04-25T00:00:00+00:00",
            "note": "legacy item sin numero_contrato_excel",
        }
        api_module._WATCH_PATH.write_text(
            json.dumps([legacy_item], ensure_ascii=False),
            encoding="utf-8",
        )

        body = client.post(
            "/watch/import-from-excel",
            json={"workbook": str(fixture_workbook_with_consecutivo)},
        ).json()

        # Debe contar la row 2 como already_recorded (no nueva appearance)
        # pero el item DEBE recibir su consecutivo via backfill.
        items = client.get("/watch").json()["items"]
        by_pid = {it["process_id"]: it for it in items}
        assert "CONTRATO-FEAB-0001-2024" in \
               by_pid["CO1.NTC.1111111"]["numero_contrato_excel"]
        # Sin duplicar appearances: la row 2 ya estaba, sigue siendo 1
        appearances_row2 = [
            a for a in by_pid["CO1.NTC.1111111"]["appearances"]
            if a["sheet"] == "FEAB 2024" and a["row"] == 2
        ]
        assert len(appearances_row2) == 1
