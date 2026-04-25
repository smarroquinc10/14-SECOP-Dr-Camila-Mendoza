"""Tests for append_process_url + delete_row — the CRM's row management.

These are tiny helpers but the Excel is the source of truth for the
Dra., so I want unit tests that catch:

1. A new row lands on the first *actually* empty row (never overwrites
   existing data).
2. The URL lands in the correct column (auto-detected).
3. Delete shifts later rows up and refuses to delete the header.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook as _load

from secop_ii.excel_io import append_process_url, delete_row


def _make_sample(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Procesos"
    ws.append(["#", "Entidad", "URL del proceso", "Objeto"])
    ws.append([1, "FEAB", "https://community.secop.gov.co/x?noticeUID=CO1.NTC.1", "Obj 1"])
    ws.append([2, "FEAB", "https://community.secop.gov.co/x?noticeUID=CO1.NTC.2", "Obj 2"])
    wb.save(path)


class TestAppendProcessUrl:
    def test_appends_to_first_empty_row(self, tmp_path):
        xlsx = tmp_path / "t.xlsx"
        _make_sample(xlsx)
        row = append_process_url(
            xlsx, "https://community.secop.gov.co/x?noticeUID=CO1.NTC.99"
        )
        assert row == 4  # header=1, then 2 existing rows, new row lands at 4

        ws = _load(xlsx).active
        assert ws.cell(row=4, column=3).value == (
            "https://community.secop.gov.co/x?noticeUID=CO1.NTC.99"
        )
        # Other columns on the new row are blank — the update pipeline fills them
        assert ws.cell(row=4, column=1).value is None
        assert ws.cell(row=4, column=4).value is None

    def test_empty_url_is_rejected(self, tmp_path):
        xlsx = tmp_path / "t.xlsx"
        _make_sample(xlsx)
        import pytest
        with pytest.raises(ValueError, match="vacía"):
            append_process_url(xlsx, "")

    def test_preserves_existing_rows(self, tmp_path):
        xlsx = tmp_path / "t.xlsx"
        _make_sample(xlsx)
        append_process_url(xlsx, "https://community.secop.gov.co/x?noticeUID=CO1.NTC.99")
        ws = _load(xlsx).active
        # Original rows untouched
        assert ws.cell(row=2, column=4).value == "Obj 1"
        assert ws.cell(row=3, column=4).value == "Obj 2"


class TestDeleteRow:
    def test_deletes_and_shifts_up(self, tmp_path):
        xlsx = tmp_path / "t.xlsx"
        _make_sample(xlsx)
        delete_row(xlsx, 2)  # kill "Obj 1"
        ws = _load(xlsx).active
        # What was row 3 ("Obj 2") is now row 2
        assert ws.cell(row=2, column=4).value == "Obj 2"
        # And row 3 is gone
        assert ws.cell(row=3, column=4).value is None

    def test_refuses_to_delete_header(self, tmp_path):
        xlsx = tmp_path / "t.xlsx"
        _make_sample(xlsx)
        import pytest
        with pytest.raises(ValueError, match="encabezado"):
            delete_row(xlsx, 1)

    def test_refuses_nonexistent_row(self, tmp_path):
        xlsx = tmp_path / "t.xlsx"
        _make_sample(xlsx)
        import pytest
        with pytest.raises(ValueError, match="no existe"):
            delete_row(xlsx, 999)
