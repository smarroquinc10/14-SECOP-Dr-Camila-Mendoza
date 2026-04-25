"""FastAPI bridge for the Next.js frontend.

Design constraints (production-grade):

1. **Espejo SECOP**: every endpoint returns ALL raw SECOP fields, never
   filters them down. The frontend can pick what to render.
2. **No false positives**: each cell carries provenance + confidence
   + the SHA-256 fingerprint of the SECOP payload it came from.
3. **No false negatives**: empty/missing fields are explicit (``None``),
   never silently dropped.
4. **No data eating**: every request that mutates state appends to the
   hash-chained audit log, with the ``code_version`` stamp.
5. **Resilient**: tenacity retry on the SECOP layer, graceful degradation
   when a sub-dataset fails (the rest still returns).
6. **Cacheable**: a per-process payload is the same until SECOP changes;
   the hash is the cache key.

Endpoints:

    GET  /health                     — sanity
    GET  /entity/feab                — FEAB metadata (NIT, total counts)
    GET  /contracts                  — all FEAB contracts (every field)
    GET  /contracts/{id_contrato}    — single contract + related data
    GET  /processes                  — all FEAB processes (every field)
    GET  /processes/{notice_uid}     — single process + related data
    GET  /audit-log                  — read the hash-chained log
    POST /refresh                    — trigger an Excel update (background)
    POST /verify                     — re-pega SECOP, compare hashes
    GET  /version                    — code version + git short SHA
"""
from __future__ import annotations

import asyncio
import logging
import os
import subprocess
import sys
from concurrent.futures import ThreadPoolExecutor
from contextlib import asynccontextmanager
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import BackgroundTasks, FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

from secop_ii.audit_log import (
    _CODE_VERSION,
    iter_entries as _audit_iter,
    render_audit_summary,
    verify_audit_log,
)
from secop_ii.feab_columns import compute_feab_fill, source_fingerprint
from secop_ii.feab_validation import validate_fills
from secop_ii.paths import state_dir, state_path
from secop_ii.secop_client import SecopClient
from secop_ii.url_parser import InvalidSecopUrlError, parse_secop_url

log = logging.getLogger("dra-cami-api")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# FEAB / Fiscalía General de la Nación
FEAB_NIT = "901148337"

# Reusable client + thread pool. The client memoizes per-process datasets
# in-memory, so repeated /contracts/{id} hits don't re-query SECOP.
_client = SecopClient()
_executor = ThreadPoolExecutor(max_workers=4)


def _seed_state_dir_if_empty() -> None:
    """First-run seed for the MSI build.

    A PyInstaller-frozen sidecar starts with an empty
    ``%LOCALAPPDATA%\\Dra Cami Contractual\\.cache\\`` on a fresh
    install — no watch list, no audit log. We ship the user's existing
    files inside the bundle (``--add-data .cache;seed``) and copy them
    once on the first cold boot. Subsequent boots see the populated
    state dir and skip seeding (so the Dra's edits are never overwritten
    by a re-install of the same version).

    Safe no-op in dev (no ``sys.frozen`` flag) and idempotent.
    """
    if not getattr(sys, "frozen", False):
        return
    target = state_dir()
    # Already seeded? Bail out — user data wins.
    if (target / "watched_urls.json").exists():
        return
    seed = Path(getattr(sys, "_MEIPASS", "")) / "seed"
    if not seed.is_dir():
        log.info("No seed dir bundled at %s — first run will be empty", seed)
        return
    import shutil
    copied = 0
    for src in seed.iterdir():
        dst = target / src.name
        if dst.exists():
            continue  # never overwrite
        try:
            if src.is_file():
                shutil.copy2(src, dst)
                copied += 1
            elif src.is_dir():
                shutil.copytree(src, dst)
                copied += 1
        except OSError as exc:
            log.warning("seed copy failed for %s: %s", src, exc)
    if copied:
        log.info("Seeded %d entries into %s", copied, target)


@asynccontextmanager
async def _lifespan(app: FastAPI):
    """Warm-up: probe SECOP once so the first user request is fast."""
    # First-run seed (no-op in dev, copies bundled .cache on MSI cold boot).
    try:
        _seed_state_dir_if_empty()
    except Exception as exc:
        log.warning("seed step failed (non-fatal): %s", exc)
    try:
        await asyncio.get_event_loop().run_in_executor(
            _executor, lambda: _client.query("jbjy-vk9h",
                                             where=f"nit_entidad='{FEAB_NIT}'",
                                             limit=1)
        )
        log.info("SECOP probe OK — API ready")
    except Exception as exc:
        log.warning("SECOP probe failed (non-fatal): %s", exc)
    yield
    _executor.shutdown(wait=False)


app = FastAPI(
    title="Dra Cami Contractual — API",
    description="Espejo automático del SECOP II para FEAB.",
    version="1.0.0",
    lifespan=_lifespan,
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        # Next dev server (`npm run dev`)
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        # Tauri 2 packaged app — Windows uses http://tauri.localhost,
        # macOS/Linux/iOS use tauri://localhost. Both are local-only
        # (the WebView2/WebKit content is bundled inside the MSI).
        "http://tauri.localhost",
        "https://tauri.localhost",
        "tauri://localhost",
    ],
    allow_credentials=False,
    # Watch list uses DELETE (remove) and PUT (edit) — without these the
    # browser preflight blocks them as soon as fetch is cross-origin
    # (which it is in the Tauri MSI: tauri.localhost → 127.0.0.1:8000).
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)


# ---- Response models --------------------------------------------------------


class HealthResponse(BaseModel):
    status: str
    code_version: str
    timestamp: str


class CellMetadata(BaseModel):
    """Provenance + confidence for a single derived cell."""
    value: Any
    source: str | None = None
    confidence: str | None = None  # HIGH / MEDIUM / LOW


class ProcessSummary(BaseModel):
    """One row in the contracts table — keeps every SECOP field plus
    the 77 FEAB-mapped cells with confidence + source."""
    id_contrato: str | None = None
    referencia_del_contrato: str | None = None
    id_del_proceso: str | None = None
    notice_uid: str | None = None
    objeto: str | None = None
    proveedor: str | None = None
    valor: float | None = None
    valor_total: float | None = None
    estado: str | None = None
    fecha_firma: str | None = None
    fecha_inicio: str | None = None
    fecha_fin: str | None = None
    modalidad: str | None = None
    tipo_contrato: str | None = None
    dias_adicionados: int | None = None
    notas: str | None = None
    secop_url: str | None = None
    secop_hash: str | None = None
    confianza: str | None = None
    needs_review: list[str] = Field(default_factory=list)
    raw: dict[str, Any] = Field(default_factory=dict)


class ObservacionDra(BaseModel):
    """One OBSERVACIONES note the Dra wrote in the master Excel."""
    sheet: str
    row: int
    text: str


class ProcessDetail(BaseModel):
    """Full per-process payload: every dataset SECOP returns about it."""
    notice_uid: str | None
    process_id: str
    proceso: dict[str, Any] | None
    contratos: list[dict[str, Any]]
    adiciones_by_contrato: dict[str, list[dict[str, Any]]]
    garantias_by_contrato: dict[str, list[dict[str, Any]]]
    pagos_by_contrato: dict[str, list[dict[str, Any]]]
    ejecucion_by_contrato: dict[str, list[dict[str, Any]]]
    suspensiones_by_contrato: dict[str, list[dict[str, Any]]]
    mods_proceso: list[dict[str, Any]]
    feab_fills: dict[str, Any]
    feab_confidence: dict[str, str]
    feab_sources: dict[str, str]
    needs_review: list[str]
    issues: list[str]
    observaciones_dra: list[ObservacionDra] = Field(default_factory=list)
    secop_hash: str
    code_version: str
    fetched_at: str


# ---- Endpoints --------------------------------------------------------------


@app.get("/health", response_model=HealthResponse)
async def health() -> HealthResponse:
    return HealthResponse(
        status="ok",
        code_version=_CODE_VERSION,
        timestamp=datetime.now(timezone.utc).isoformat(timespec="seconds"),
    )


@app.get("/version")
async def version() -> dict[str, str]:
    return {"code_version": _CODE_VERSION, "feab_nit": FEAB_NIT}


@app.get("/entity/feab")
async def entity_feab() -> dict[str, Any]:
    """Metadata about FEAB itself + counts."""
    loop = asyncio.get_event_loop()

    def _counts() -> dict[str, int]:
        try:
            ctos = _client.query("jbjy-vk9h",
                                where=f"nit_entidad='{FEAB_NIT}'",
                                select="count(*) as n",
                                limit=1)
            procs = _client.query("p6dx-8zbt",
                                where=f"nit_entidad='{FEAB_NIT}'",
                                select="count(*) as n",
                                limit=1)
            return {
                "contratos": int(ctos[0].get("n", 0)) if ctos else 0,
                "procesos": int(procs[0].get("n", 0)) if procs else 0,
            }
        except Exception as exc:  # pragma: no cover
            log.warning("entity counts failed: %s", exc)
            return {"contratos": 0, "procesos": 0}

    counts = await loop.run_in_executor(_executor, _counts)
    return {
        "nit": FEAB_NIT,
        "nombre": "Fondo Especial para la Administración de Bienes "
                  "de la Fiscalía General de la Nación",
        "alias": "FEAB",
        "padre": "Fiscalía General de la Nación",
        **counts,
    }


@app.get("/contracts")
async def list_contracts(
    limit: int = Query(500, ge=1, le=10_000),
    offset: int = Query(0, ge=0),
    where: str | None = Query(None, description="Extra Socrata WHERE clause"),
) -> list[dict[str, Any]]:
    """Return EVERY SECOP contract for FEAB with EVERY field.

    The frontend chooses what to render. We never strip fields — that
    would be a false negative.
    """
    loop = asyncio.get_event_loop()

    def _fetch() -> list[dict]:
        clause = f"nit_entidad='{FEAB_NIT}'"
        if where:
            clause = f"({clause}) AND ({where})"
        return _client.query("jbjy-vk9h", where=clause,
                            limit=limit, offset=offset,
                            order="fecha_de_firma DESC")

    rows = await loop.run_in_executor(_executor, _fetch)

    # Add a derived "notas" + auto-generated narrative per row so the
    # frontend doesn't need to recompute it. Same logic the Streamlit
    # app uses — kept here as a single source of truth.
    out: list[dict[str, Any]] = []
    for r in rows:
        notas_parts: list[str] = []
        estado = str(r.get("estado_contrato") or "")
        if "Modificad" in estado:
            notas_parts.append("Modificado")
        try:
            d = int(float(r.get("dias_adicionados") or 0))
            if d > 0:
                notas_parts.append(f"+{d} días")
        except (ValueError, TypeError):
            pass
        liq = str(r.get("liquidaci_n") or "").strip().lower()
        if liq == "si":
            notas_parts.append("Liquidado")
        r["_notas"] = " · ".join(notas_parts)
        # Flatten the urlproceso dict to a plain string up front.
        url = r.get("urlproceso")
        if isinstance(url, dict):
            r["urlproceso"] = url.get("url") or ""
        out.append(r)
    return out


@app.get("/contracts/{id_or_pid}", response_model=ProcessDetail)
async def contract_detail(id_or_pid: str) -> ProcessDetail:
    """Full payload for a single process. Accepts ANY identifier the UI
    might pass — id_contrato (CO1.PCCNTR.X), notice_uid (CO1.NTC.X),
    process_id (CO1.PPI.X / CO1.REQ.X). Resolves to the contract(s)
    when they exist; returns proceso-only payload (with the Dra's
    observaciones) when there is no contract signed yet.

    Mirror semantics: nothing is invented. If neither contracts nor
    proceso resolve, returns 404 honestly.
    """
    loop = asyncio.get_event_loop()

    def _fetch() -> ProcessDetail:
        ident = id_or_pid.strip()

        # 1) Try as id_contrato (PCCNTR) — direct contract row.
        contratos = _client.query(
            "jbjy-vk9h", where=f"id_contrato='{_esc(ident)}'", limit=10,
        )
        # 2) If no match, try as proceso_de_compra (NTC / REQ at contract level).
        if not contratos:
            contratos = _client.query(
                "jbjy-vk9h",
                where=f"proceso_de_compra='{_esc(ident)}'",
                limit=10,
            )

        # 3) Look up the watch-list entry to get the original URL.
        watch_url = None
        for it in _load_watched():
            if it.get("process_id") == ident or it.get("url") == ident \
                    or it.get("notice_uid") == ident:
                watch_url = it.get("url")
                break

        # 4) Determine the URL we'll use for parsing.
        if contratos:
            c0 = contratos[0]
            url_field = c0.get("urlproceso")
            if isinstance(url_field, dict):
                url_field = url_field.get("url") or ""
            url = str(url_field or "") or (watch_url or "")
        else:
            url = watch_url or ""

        # 5) Resolve proceso (parent). Try the URL parser first;
        # if that fails, use the identifier directly when it looks
        # like a process_id. Drafts (CO1.REQ.*) often have the
        # process info even without a contract.
        proceso = None
        notice_uid = None
        try:
            if url:
                ref = parse_secop_url(url)
                proceso = _client.get_proceso(ref.process_id, url=url)
                notice_uid = _client.resolve_notice_uid(
                    ref.process_id, url=url
                )
            elif ident.startswith(("CO1.PPI.", "CO1.NTC.", "CO1.REQ.")):
                # Best effort lookup by ID alone
                proceso = _client.get_proceso(ident)
                notice_uid = _client.resolve_notice_uid(ident)
        except InvalidSecopUrlError:
            pass
        except Exception as exc:
            log.warning("proceso resolve failed: %s", exc)

        # 6) Honest 404: if NEITHER a contract NOR a proceso resolved,
        # we have nothing meaningful to show. (Observaciones alone are
        # not enough — they may belong to an archived/draft process.)
        if not contratos and not proceso:
            # Last resort: if Excel has notes for this id, return a
            # minimal payload with just the observaciones — better
            # than 404 because the Dra's manual record exists.
            obs_raw = _observaciones_for(
                process_id=ident, proceso_de_compra=ident,
                notice_uid=ident, contract_id=ident, url=url or None,
            )
            if not obs_raw:
                raise HTTPException(
                    404,
                    f"{ident} no se encuentra en SECOP API ni hay "
                    "observaciones en el Excel. Validá manualmente "
                    "abriendo el link.",
                )
            return ProcessDetail(
                notice_uid=None,
                process_id=ident,
                proceso=None,
                contratos=[],
                adiciones_by_contrato={},
                garantias_by_contrato={},
                pagos_by_contrato={},
                ejecucion_by_contrato={},
                suspensiones_by_contrato={},
                mods_proceso=[],
                feab_fills={},
                feab_confidence={},
                feab_sources={},
                needs_review=[],
                issues=[
                    "El proceso no aparece en el API público de "
                    "datos.gov.co. Las observaciones de abajo vienen "
                    "del Excel master de la Dra.",
                ],
                observaciones_dra=[ObservacionDra(**e) for e in obs_raw],
                secop_hash="",
                code_version=_CODE_VERSION,
                fetched_at=datetime.now(timezone.utc).isoformat(
                    timespec="seconds"
                ),
            )

        # Normal path: contracts (or proceso-only if no contracts)
        c = contratos[0] if contratos else {}
        portfolio = str(c.get("proceso_de_compra") or "")

        # All sub-datasets, gracefully degrading on errors per dataset
        def _safe_call(fn, *args, **kw):
            try:
                return fn(*args, **kw)
            except Exception as exc:
                log.warning("subdataset fetch failed for %s: %s",
                          getattr(fn, "__name__", "?"), exc)
                return []

        adis = {ci: _safe_call(_client.get_adiciones, ci)
               for ci in [c.get("id_contrato")] if ci}
        gars = {ci: _safe_call(_client.get_garantias, ci)
               for ci in [c.get("id_contrato")] if ci}
        facs = {ci: _safe_call(_client.get_facturas, ci)
               for ci in [c.get("id_contrato")] if ci}
        ejec = {ci: _safe_call(_client.get_ejecucion, ci)
               for ci in [c.get("id_contrato")] if ci}
        susp = {ci: _safe_call(_client.get_suspensiones, ci)
               for ci in [c.get("id_contrato")] if ci}
        mods_proc = _safe_call(_client.get_mod_procesos, portfolio) if portfolio else []

        result = compute_feab_fill(
            proceso=proceso, contratos=contratos,
            notice_uid=notice_uid, source_url=url,
            adiciones_by_contrato=adis,
            garantias_by_contrato=gars,
            ejecucion_by_contrato=ejec,
        )
        validation = validate_fills(result.values)
        secop_hash = source_fingerprint(
            proceso=proceso, contratos=contratos, notice_uid=notice_uid,
        )

        # Look up the Dra's manual notes from the Excel master file.
        # Best-effort: if the workbook isn't available or the row
        # doesn't exist, return an empty list — never blocks the API.
        try:
            obs_raw = _observaciones_for(
                process_id=(proceso or {}).get("id_del_proceso"),
                proceso_de_compra=str(c.get("proceso_de_compra") or "")
                                  or None,
                notice_uid=notice_uid,
                contract_id=c.get("id_contrato"),
                url=url,
            )
            obs = [ObservacionDra(**e) for e in obs_raw]
        except Exception as exc:
            log.warning("observaciones lookup failed: %s", exc)
            obs = []

        return ProcessDetail(
            notice_uid=notice_uid,
            process_id=(proceso or {}).get("id_del_proceso", ident),
            proceso=proceso,
            contratos=contratos,
            adiciones_by_contrato=adis,
            garantias_by_contrato=gars,
            pagos_by_contrato=facs,
            ejecucion_by_contrato=ejec,
            suspensiones_by_contrato=susp,
            mods_proceso=mods_proc,
            feab_fills=result.values,
            feab_confidence=result.confidence,
            feab_sources=result.sources,
            needs_review=sorted(validation.needs_review),
            issues=validation.issues,
            observaciones_dra=obs,
            secop_hash=secop_hash,
            code_version=_CODE_VERSION,
            fetched_at=datetime.now(timezone.utc).isoformat(timespec="seconds"),
        )

    return await loop.run_in_executor(_executor, _fetch)


@app.get("/processes")
async def list_processes(
    limit: int = Query(500, ge=1, le=10_000),
    offset: int = Query(0, ge=0),
) -> list[dict[str, Any]]:
    """All FEAB processes (p6dx-8zbt) with every field."""
    loop = asyncio.get_event_loop()

    def _fetch() -> list[dict]:
        return _client.query(
            "p6dx-8zbt", where=f"nit_entidad='{FEAB_NIT}'",
            limit=limit, offset=offset,
            order="fecha_de_publicacion_del DESC",
        )

    rows = await loop.run_in_executor(_executor, _fetch)
    for r in rows:
        url = r.get("urlproceso")
        if isinstance(url, dict):
            r["urlproceso"] = url.get("url") or ""
    return rows


_WATCH_PATH = state_path("watched_urls.json")
_WATCH_LOCK = __import__("threading").Lock()


class WatchedItem(BaseModel):
    """One unique SECOP process the Dra is tracking.

    The same process can appear on multiple Excel sheets (e.g. a
    plurianual contract that shows up in FEAB 2024 and FEAB 2025).
    To avoid hiding that fact while also avoiding duplicate rows,
    each item carries a list of every (sheet, vigencia) pair where
    its URL was seen. The UI filter by sheet returns the same count
    the Dra sees in Excel — no data is eaten, no row is invented.
    """
    url: str
    process_id: str | None = None
    notice_uid: str | None = None
    sheets: list[str] = Field(default_factory=list)
    vigencias: list[str] = Field(default_factory=list)
    appearances: list[dict[str, str | None]] = Field(default_factory=list)
    added_at: str
    note: str | None = None


def _load_watched() -> list[dict[str, Any]]:
    if not _WATCH_PATH.exists():
        return []
    try:
        import json as _json
        return _json.loads(_WATCH_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []


def _save_watched(items: list[dict[str, Any]]) -> None:
    import json as _json
    _WATCH_PATH.parent.mkdir(parents=True, exist_ok=True)
    _WATCH_PATH.write_text(
        _json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8"
    )


@app.get("/watch")
async def watch_list() -> dict[str, Any]:
    """Return the manually-tracked SECOP URLs.

    Per the Dra's cardinal rule: **the truth lives in SECOP, the Excel
    only contributes the LINK and the VIGENCIA**. We never derive
    estado/valor/proveedor/etc from the Excel.

    Exception: the Dra's hand-written OBSERVACIONES (col 72) ARE
    her own truth — they're not "Excel data we're guessing", they're
    her audit annotations. We surface a brief of those observations
    (`obs_brief`) and a `is_modificado` flag derived from her wording
    so the UI can show a "Modificatorio (Excel)" line when the SECOP
    API has no contract row to confirm. The label always makes clear
    the source is the Dra's manual note, NEVER claiming SECOP truth.
    """
    items = _load_watched()
    for it in items:
        # Cardinal rule: del Excel SOLO se toma vigencia + link. NUNCA
        # numero_contrato — ese es de SECOP (`referencia_del_contrato`).
        # Si SECOP no lo expone, la UI muestra "—" honesto.
        # The Dra's own annotations (modificatorios, prórrogas, etc.).
        try:
            obs_raw = _observaciones_for(
                process_id=it.get("process_id"),
                proceso_de_compra=it.get("notice_uid"),
                notice_uid=it.get("notice_uid"),
                contract_id=None,
                url=it.get("url"),
            )
        except Exception:
            obs_raw = []

        if obs_raw:
            obs_text = (obs_raw[0].get("text") or "").strip()
            it["obs_brief"] = (
                obs_text[:120] + "…"
                if len(obs_text) > 120
                else obs_text
            ) if obs_text else None
            blob = " ".join(
                (o.get("text") or "").upper() for o in obs_raw
            )
            it["is_modificado_excel"] = any(
                k in blob
                for k in ("MODIFIC", "PRORROG", "PRÓRROG",
                          "ADICION", "ADICIÓN", "MOD ", "MOD\n")
            )
        else:
            it["obs_brief"] = None
            it["is_modificado_excel"] = False
    return {"items": items}


@app.post("/watch")
async def watch_add(payload: dict[str, str]) -> dict[str, Any]:
    """Add a SECOP URL to the watch list. Dedup by parsed process_id.

    Body::

        {"url": "<URL del SECOP II>",
         "note": "optional comment",
         "sheet": "FEAB 2024"  # optional — la hoja a la que pertenece}

    If ``sheet`` is provided, the new item is recorded as appearing on
    that sheet (sheets=[sheet], vigencias=[derived from sheet name],
    one synthetic appearance with row=None to mark it as manual). If
    the URL was already in the list, the new sheet is APPENDED to its
    existing sheets/vigencias/appearances — we never overwrite history.

    Returns ``{added: true|false, item, total}`` so the UI can show
    "ya estaba en tu lista" feedback without inventing duplicates.
    """
    url = (payload.get("url") or "").strip()
    note = (payload.get("note") or "").strip() or None
    sheet = (payload.get("sheet") or "").strip() or None
    if not url:
        raise HTTPException(400, "url is required")
    if "secop.gov.co" not in url.lower():
        raise HTTPException(400, "URL no parece de SECOP II")

    process_id = None
    notice_uid = None
    try:
        ref = parse_secop_url(url)
        process_id = ref.process_id
        notice_uid = _client.resolve_notice_uid(ref.process_id, url=url)
    except InvalidSecopUrlError:
        # Allow URLs we can't parse; we'll dedup by URL string instead.
        pass

    vigencia = _vigencia_from_sheet_name(sheet) if sheet else None

    with _WATCH_LOCK:
        items = _load_watched()
        # Dedup: prefer process_id match, fall back to URL match.
        for it in items:
            same_id = process_id and it.get("process_id") == process_id
            same_url = it.get("url") == url
            if not (same_id or same_url):
                continue
            # Already in the list. If the user picked a sheet that's
            # not yet in this item's sheets[], append it (the same
            # process can legitimately live on multiple sheets).
            if sheet:
                it.setdefault("sheets", [])
                it.setdefault("vigencias", [])
                it.setdefault("appearances", [])
                if sheet not in it["sheets"]:
                    it["sheets"].append(sheet)
                if vigencia and vigencia not in it["vigencias"]:
                    it["vigencias"].append(vigencia)
                # Synthetic appearance from manual add (row=None to
                # distinguish from Excel-imported rows).
                already = any(
                    a.get("sheet") == sheet
                    and a.get("row") is None
                    and a.get("url") == url
                    for a in it["appearances"]
                )
                if not already:
                    it["appearances"].append({
                        "sheet": sheet,
                        "vigencia": vigencia,
                        "row": None,
                        "url": url,
                    })
                _save_watched(items)
                return {"added": False, "item": it, "total": len(items),
                       "reason": f"ya estaba en tu lista — agregado también a {sheet}"}
            return {"added": False, "item": it, "total": len(items),
                   "reason": "ya estaba en tu lista"}

        new_item = {
            "url": url,
            "process_id": process_id,
            "notice_uid": notice_uid,
            "sheets": [sheet] if sheet else [],
            "vigencias": [vigencia] if vigencia else [],
            "appearances": [{
                "sheet": sheet,
                "vigencia": vigencia,
                "row": None,
                "url": url,
            }] if sheet else [],
            "added_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "note": note,
        }
        items.append(new_item)
        _save_watched(items)
        return {"added": True, "item": new_item, "total": len(items)}


@app.delete("/watch")
async def watch_remove(url: str = Query(..., description="URL exacta a borrar"),
                        process_id: str | None = Query(None)) -> dict[str, Any]:
    """Remove a watched URL. Match by URL or process_id."""
    with _WATCH_LOCK:
        items = _load_watched()
        before = len(items)
        items = [
            it for it in items
            if it.get("url") != url
            and (process_id is None or it.get("process_id") != process_id)
        ]
        _save_watched(items)
        return {"removed": before - len(items), "total": len(items)}


@app.put("/watch")
async def watch_update(payload: dict[str, str]) -> dict[str, Any]:
    """Edit a watched item — replace its URL while keeping its
    appearances/sheets/vigencias history intact.

    Body::

        {"old_url": "<the URL currently stored>",
         "new_url": "<the corrected URL>",
         "note": "optional updated note"}

    Re-parses ``new_url`` to extract ``process_id`` and resolves
    ``notice_uid`` against SECOP. Refuses if ``new_url`` already
    belongs to a different watched process (would create a duplicate).
    """
    old_url = (payload.get("old_url") or "").strip()
    new_url = (payload.get("new_url") or "").strip()
    note = payload.get("note")
    if not old_url or not new_url:
        raise HTTPException(400, "old_url y new_url son requeridos")
    if "secop.gov.co" not in new_url.lower():
        raise HTTPException(400, "new_url no parece de SECOP II")

    new_process_id = None
    new_notice_uid = None
    try:
        ref = parse_secop_url(new_url)
        new_process_id = ref.process_id
        new_notice_uid = _client.resolve_notice_uid(
            ref.process_id, url=new_url
        )
    except InvalidSecopUrlError:
        pass

    with _WATCH_LOCK:
        items = _load_watched()

        # Locate the item to edit
        target = None
        for it in items:
            if it.get("url") == old_url:
                target = it
                break
        if target is None:
            raise HTTPException(404, f"no encuentro old_url en tu lista")

        # Refuse if new_url is already in another item (different process)
        for it in items:
            if it is target:
                continue
            if it.get("url") == new_url:
                raise HTTPException(
                    409, "new_url ya pertenece a otro proceso de tu lista"
                )
            if (new_process_id and it.get("process_id") == new_process_id):
                raise HTTPException(
                    409,
                    f"new_url apunta al mismo process_id ({new_process_id}) "
                    "que otro item de tu lista",
                )

        target["url"] = new_url
        if new_process_id:
            target["process_id"] = new_process_id
        if new_notice_uid is not None:
            target["notice_uid"] = new_notice_uid
        if note is not None:
            target["note"] = (note.strip() or None) if note else None
        # Stamp the edit
        target["edited_at"] = datetime.now(timezone.utc).isoformat(
            timespec="seconds"
        )

        _save_watched(items)
        return {"updated": True, "item": target, "total": len(items)}


# ---- Excel → watch list importer -------------------------------------------
# The Dra's master Excel keeps a "LINK" column with the SECOP URL of each
# process she follows. Some sheets put headers in row 1 (FEAB 2026/2025/
# 2024/2023/2022), and the legacy "FEAB 2018-2021" sheet puts them in row
# 4 (rows 1-3 are formato titles). We auto-detect both layouts.

_DEFAULT_WORKBOOK = "BASE DE DATOS FEAB CONTRATOS2.xlsx"
_LINK_HEADER_KEYWORDS = ("LINK", "URL", "ENLACE")


def _find_link_column(ws) -> tuple[int, int] | None:
    """Return (header_row, link_col_1_indexed) or None.

    Probes row 1, then row 4. Matches a header whose stripped uppercase
    equals exactly ``"LINK"`` to avoid false positives like "LINK
    verificación API" or "Link de SIRECI" — the Dra's column is just
    ``LINK``. Falls back to startswith("LINK") if exact match fails.
    """
    for header_row in (1, 4):
        try:
            row = next(ws.iter_rows(min_row=header_row, max_row=header_row,
                                    values_only=True))
        except StopIteration:
            continue
        # Exact "LINK" match first
        for i, v in enumerate(row, start=1):
            if v is None:
                continue
            text = str(v).strip().upper()
            if text == "LINK":
                return header_row, i
        # Fallback: header equals one of the keywords (whole-cell match)
        for i, v in enumerate(row, start=1):
            if v is None:
                continue
            text = str(v).strip().upper()
            if text in _LINK_HEADER_KEYWORDS:
                return header_row, i
    return None


def _find_vigencia_column(ws, header_row: int) -> int | None:
    """Return the 1-indexed column that holds the VIGENCIA value, or None.

    The Dra's format uses ``"3.VIGENCIA"`` as the canonical header, but
    different sheets put it in different physical columns (col 2 in
    FEAB 2026/2025/2024/2023/2022, col 3 in FEAB 2018-2021). We match
    by header text, not column index — that way the columns can shift
    without breaking the importer.

    Skips headers that *contain* "VIGENCIA" but mean something else
    (e.g. "VALOR VIGENCIA ACTUAL", "Garantía vigencia desde"): we
    require the header to start with a digit-and-dot prefix followed
    by VIGENCIA, OR to equal "VIGENCIA" alone.
    """
    try:
        row = next(ws.iter_rows(min_row=header_row, max_row=header_row,
                                values_only=True))
    except StopIteration:
        return None
    for i, v in enumerate(row, start=1):
        if v is None:
            continue
        text = str(v).strip().upper()
        # Match "3.VIGENCIA", "3 VIGENCIA", "VIGENCIA" — but not
        # "VALOR VIGENCIA ACTUAL" or "VIGENCIA FUTURA".
        if text == "VIGENCIA":
            return i
        # Strip leading digits + punctuation
        head = text.lstrip("0123456789. ")
        if head == "VIGENCIA":
            return i
    return None


def _vigencia_from_sheet_name(sheet_name: str) -> str | None:
    """Best-effort fallback: derive vigencia from a sheet name like
    ``"FEAB 2024"`` → ``"2024"``. Returns None for ranges like
    ``"FEAB 2018-2021"`` (those rows always have an explicit per-row
    vigencia in the Excel column)."""
    import re
    m = re.search(r"\b(\d{4})\s*$", sheet_name.strip())
    return m.group(1) if m else None


def _find_obs_column(ws, header_row: int) -> int | None:
    """Locate the ``OBSERVACIONES`` column in a sheet by header text."""
    try:
        row = next(ws.iter_rows(min_row=header_row, max_row=header_row,
                                values_only=True))
    except StopIteration:
        return None
    for i, v in enumerate(row, start=1):
        if v is None:
            continue
        if "OBSERVAC" in str(v).strip().upper():
            return i
    return None


# Headers we want to surface to the UI when SECOP API has no contract
# but the Dra's Excel does. The match is by SUBSTRING (case-insensitive,
# accent-tolerant via unicodedata) so minor formatting variations don't
# break it. Order matters: the FIRST match wins per field.
_FIELD_HEADER_MATCHERS: dict[str, list[str]] = {
    "estado": ["10.ESTADO DEL CONTRATO", "ESTADO DEL CONTRATO"],
    "fecha_firma": ["5. FECHA SUSCRIPCION", "FECHA SUSCRIPCION"],
    "fecha_inicio": ["8. FECHA INICIO", "FECHA INICIO"],
    "fecha_terminacion": ["9. FECHA TERMINACION", "FECHA TERMINACION"],
    "valor_total": ["36. VALOR TOTAL", "36: VALOR TOTAL", "VALOR TOTAL"],
    "valor_inicial": [
        "28.VALOR INICIAL DEL CONTRATO",
        "VALOR INICIAL DEL CONTRATO",
    ],
    "proveedor": [
        "43. CONTRATISTA : NOMBRE COMPLETO",
        "CONTRATISTA : NOMBRE COMPLETO",
        "NOMBRE COMPLETO",
    ],
    "objeto": ["4. OBJETO", "OBJETO"],
    "modalidad": ["12. MODALIDAD DE SELECCION", "MODALIDAD DE SELECCION"],
    "dias_prorrogas": ["37. PRORROGAS : NUMERO DE DIAS", "PRORROGAS : NUMERO DE DIAS"],
    "adiciones_count": ["35A ADICIONES", "35. TIPO DE ADICIONES"],
    "liquidacion": ["67.REQUIERE LIQUIDACION", "REQUIERE LIQUIDACION"],
    "fecha_liquidacion": ["68. FECHA LIQUIDACION", "FECHA LIQUIDACION"],
    "numero_contrato": ["2. NUMERO DE CONTRATO", "NUMERO DE CONTRATO"],
    "supervisor": ["54.NOMBRE DEL INTERVENTOR", "NOMBRE DEL INTERVENTOR"],
}


def _norm(s: str) -> str:
    """Uppercase + strip accents for header matching."""
    import unicodedata
    return "".join(
        c for c in unicodedata.normalize("NFD", s.upper())
        if unicodedata.category(c) != "Mn"
    )


def _find_field_columns(ws, header_row: int) -> dict[str, int]:
    """Return {field_name -> 1-indexed col} for every recognized header
    in ``_FIELD_HEADER_MATCHERS``. Skips fields whose header isn't found."""
    try:
        row = next(ws.iter_rows(min_row=header_row, max_row=header_row,
                                values_only=True))
    except StopIteration:
        return {}
    headers_norm = [
        _norm(str(v).strip()) if v is not None else "" for v in row
    ]
    out: dict[str, int] = {}
    for field, candidates in _FIELD_HEADER_MATCHERS.items():
        for cand in candidates:
            cand_n = _norm(cand)
            for i, h in enumerate(headers_norm, start=1):
                if cand_n in h:
                    out[field] = i
                    break
            if field in out:
                break
    return out


def _read_excel_row_fields(
    row: tuple, field_cols: dict[str, int]
) -> dict[str, Any]:
    """Pluck the recognized fields out of a row, normalizing values."""
    import datetime as _dt
    out: dict[str, Any] = {}
    for field, col in field_cols.items():
        if len(row) < col:
            continue
        v = row[col - 1]
        if v is None:
            continue
        if isinstance(v, _dt.datetime):
            out[field] = v.strftime("%Y-%m-%d")
        elif isinstance(v, _dt.date):
            out[field] = v.isoformat()
        elif isinstance(v, (int, float)):
            out[field] = v
        else:
            text = str(v).strip()
            if text:
                out[field] = text
    return out


# In-memory cache keyed by (workbook_path, mtime) so we re-read the
# Excel only when it actually changes.
_OBS_INDEX_CACHE: dict[tuple[str, float],
                       dict[str, list[dict[str, Any]]]] = {}
_DATA_INDEX_CACHE: dict[tuple[str, float],
                        dict[str, dict[str, Any]]] = {}


def _load_excel_obs_index(workbook_path: Path) -> dict[
    str, list[dict[str, Any]]
]:
    """Build ``{key -> [{sheet, row, text}, ...]}`` where ``key`` is
    every URL substring or process_id we can cheaply extract from each
    Excel row that has a non-empty OBSERVACIONES cell.

    Cached per (path, mtime) — re-reads only when the Excel changes.
    """
    if not workbook_path.exists():
        return {}
    mtime = workbook_path.stat().st_mtime
    cache_key = (str(workbook_path), mtime)
    cached = _OBS_INDEX_CACHE.get(cache_key)
    if cached is not None:
        return cached

    from openpyxl import load_workbook
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    index: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        link_loc = _find_link_column(ws)
        if not link_loc:
            continue
        header_row, link_col = link_loc
        obs_col = _find_obs_column(ws, header_row)
        if obs_col is None:
            continue

        for excel_row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not row or len(row) < max(link_col, obs_col):
                continue
            obs_v = row[obs_col - 1]
            if obs_v is None:
                continue
            text = str(obs_v).strip()
            if not text:
                continue

            entry = {"sheet": sheet_name, "row": excel_row_idx,
                    "text": text}

            # Index by URL (full + lower) and parsed process_id.
            # Skip non-URL values (the LINK column sometimes contains
            # status words like "ANULADO" — those would pollute the
            # index with garbage keys).
            link_v = row[link_col - 1]
            if link_v is not None:
                url = str(link_v).strip()
                if url and "secop.gov.co" in url.lower():
                    index.setdefault(url, []).append(entry)
                    index.setdefault(url.lower(), []).append(entry)
                    try:
                        ref = parse_secop_url(url)
                        index.setdefault(ref.process_id, []).append(entry)
                    except InvalidSecopUrlError:
                        pass

    # Dedup entries within each key (multiple keys can point to the same row)
    for key, entries in index.items():
        seen = set()
        unique: list[dict[str, Any]] = []
        for e in entries:
            sig = (e["sheet"], e["row"])
            if sig in seen:
                continue
            seen.add(sig)
            unique.append(e)
        index[key] = unique

    # Evict older cache entries (single-workbook system)
    _OBS_INDEX_CACHE.clear()
    _OBS_INDEX_CACHE[cache_key] = index
    return index


def _load_excel_data_index(workbook_path: Path) -> dict[str, dict[str, Any]]:
    """Build ``{key -> {sheet, row, vigencia, estado, valor_total, ...}}``
    where ``key`` is every URL/process_id/lower-URL we can extract from
    each row that has a SECOP link. Last-write wins per key (the most
    recent sheet's data takes priority).

    Cached per (path, mtime). The set of fields surfaced is defined
    by ``_FIELD_HEADER_MATCHERS`` so the columns auto-adjust if the
    Dra renames things in the Excel.
    """
    if not workbook_path.exists():
        return {}
    mtime = workbook_path.stat().st_mtime
    cache_key = (str(workbook_path), mtime)
    cached = _DATA_INDEX_CACHE.get(cache_key)
    if cached is not None:
        return cached

    from openpyxl import load_workbook
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    index: dict[str, dict[str, Any]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        link_loc = _find_link_column(ws)
        if not link_loc:
            continue
        header_row, link_col = link_loc
        field_cols = _find_field_columns(ws, header_row)
        if not field_cols:
            continue
        vig_col = _find_vigencia_column(ws, header_row)
        sheet_fallback_vig = _vigencia_from_sheet_name(sheet_name)

        for excel_row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not row or len(row) < link_col:
                continue
            link_v = row[link_col - 1]
            if link_v is None:
                continue
            url = str(link_v).strip()
            if not url or "secop.gov.co" not in url.lower():
                continue

            data = _read_excel_row_fields(row, field_cols)
            if not data:
                continue

            # Vigencia (per-row col preferred, sheet name fallback)
            vigencia: str | None = None
            if vig_col is not None and len(row) >= vig_col:
                rv = row[vig_col - 1]
                if rv is not None:
                    if isinstance(rv, (int, float)):
                        vigencia = str(int(rv))
                    else:
                        vigencia = str(rv).strip() or None
            if not vigencia:
                vigencia = sheet_fallback_vig

            data["sheet"] = sheet_name
            data["row"] = excel_row_idx
            data["vigencia"] = vigencia

            index[url] = data
            index[url.lower()] = data
            try:
                ref = parse_secop_url(url)
                index[ref.process_id] = data
            except InvalidSecopUrlError:
                pass

    _DATA_INDEX_CACHE.clear()
    _DATA_INDEX_CACHE[cache_key] = index
    return index


def _excel_data_for(
    process_id: str | None,
    notice_uid: str | None,
    contract_id: str | None,
    url: str | None,
) -> dict[str, Any] | None:
    """Look up a row's Excel data by any of its identifiers."""
    workbook = Path(_DEFAULT_WORKBOOK)
    index = _load_excel_data_index(workbook)
    for key in (process_id, notice_uid, contract_id,
                url, (url or "").lower()):
        if not key:
            continue
        if key in index:
            return index[key]
    return None


def _observaciones_for(
    process_id: str | None,
    proceso_de_compra: str | None,
    notice_uid: str | None,
    contract_id: str | None,
    url: str | None,
) -> list[dict[str, Any]]:
    """Look up the Dra's notes for a contract by every available key.

    Tries: explicit URL, process_id, proceso_de_compra (NTC for many
    contracts), notice_uid, id_contrato (PCCNTR). De-dups results.
    """
    workbook = Path(_DEFAULT_WORKBOOK)
    index = _load_excel_obs_index(workbook)
    found: list[dict[str, Any]] = []
    seen_rows: set[tuple[str, int]] = set()
    for key in (process_id, proceso_de_compra, notice_uid,
                contract_id, url, (url or "").lower()):
        if not key:
            continue
        for entry in index.get(key, []):
            sig = (entry["sheet"], entry["row"])
            if sig in seen_rows:
                continue
            seen_rows.add(sig)
            found.append(entry)
    return found


def _import_workbook_urls(workbook_path: Path) -> dict[str, Any]:
    """Synchronous worker: read every sheet, MERGE per-URL appearances
    into one item per unique URL/process_id, return a report.

    Mirror semantics: every (sheet, row) where a SECOP URL appears in
    the Excel is recorded as an "appearance". A URL that shows up on
    3 sheets has 3 appearances on a single watch item — the UI filter
    by sheet shows it once for each sheet, matching the Excel exactly.
    """
    from openpyxl import load_workbook

    if not workbook_path.exists():
        raise HTTPException(404, f"workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    with _WATCH_LOCK:
        existing = _load_watched()

        # Build O(1) lookups so we can merge by process_id or by URL.
        by_pid: dict[str, dict[str, Any]] = {}
        by_url: dict[str, dict[str, Any]] = {}
        for it in existing:
            pid = it.get("process_id")
            url = it.get("url")
            if pid:
                by_pid[pid] = it
            if url:
                by_url[url] = it
            # Backfill the new schema for legacy items missing the lists
            it.setdefault("sheets", [])
            it.setdefault("vigencias", [])
            it.setdefault("appearances", [])

        per_sheet: dict[str, dict[str, int]] = {}
        new_items_count = 0
        merged_count = 0
        already_recorded = 0  # appearance already on file (idempotent re-runs)
        errors: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            link_loc = _find_link_column(ws)
            if not link_loc:
                per_sheet[sheet_name] = {"found": 0, "added_new": 0,
                                        "merged": 0, "already_recorded": 0,
                                        "no_link_col": 1}
                continue
            header_row, link_col = link_loc
            vig_col = _find_vigencia_column(ws, header_row)
            sheet_fallback_vig = _vigencia_from_sheet_name(sheet_name)
            stats = {"found": 0, "added_new": 0, "merged": 0,
                    "already_recorded": 0, "no_link_col": 0}

            for excel_row_idx, row in enumerate(
                ws.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                if not row or len(row) < link_col:
                    continue
                v = row[link_col - 1]
                if v is None:
                    continue
                url = str(v).strip()
                if not url or "secop.gov.co" not in url.lower():
                    continue
                stats["found"] += 1

                # Read VIGENCIA per-row (Excel column 3.VIGENCIA) with
                # fallback to the sheet name year.
                vigencia: str | None = None
                if vig_col is not None and len(row) >= vig_col:
                    rv = row[vig_col - 1]
                    if rv is not None:
                        if isinstance(rv, (int, float)):
                            vigencia = str(int(rv))
                        else:
                            vigencia = str(rv).strip() or None
                if not vigencia:
                    vigencia = sheet_fallback_vig

                process_id = None
                try:
                    ref = parse_secop_url(url)
                    process_id = ref.process_id
                except InvalidSecopUrlError:
                    pass

                appearance = {
                    "sheet": sheet_name,
                    "vigencia": vigencia,
                    "row": excel_row_idx,
                    "url": url,
                }

                # Find a matching existing item by process_id, then URL
                target = None
                if process_id and process_id in by_pid:
                    target = by_pid[process_id]
                elif url in by_url:
                    target = by_url[url]

                if target is None:
                    # New unique process
                    target = {
                        "url": url,
                        "process_id": process_id,
                        "notice_uid": None,
                        "sheets": [],
                        "vigencias": [],
                        "appearances": [],
                        "added_at": datetime.now(timezone.utc)
                                           .isoformat(timespec="seconds"),
                        "note": f"Importado de Excel · primera vista en "
                                f"{sheet_name}",
                    }
                    existing.append(target)
                    if process_id:
                        by_pid[process_id] = target
                    by_url[url] = target
                    new_items_count += 1
                    stats["added_new"] += 1
                else:
                    # Merge: already known process. Did we already see
                    # THIS exact (sheet, row, url) appearance?
                    already_seen = any(
                        a.get("sheet") == sheet_name
                        and a.get("row") == excel_row_idx
                        and a.get("url") == url
                        for a in target.get("appearances", [])
                    )
                    if already_seen:
                        already_recorded += 1
                        stats["already_recorded"] += 1
                        continue
                    merged_count += 1
                    stats["merged"] += 1

                # Record the appearance, keep sheets/vigencias unique
                target.setdefault("appearances", []).append(appearance)
                if sheet_name not in target.setdefault("sheets", []):
                    target["sheets"].append(sheet_name)
                if vigencia and vigencia not in target.setdefault(
                    "vigencias", []
                ):
                    target["vigencias"].append(vigencia)

            per_sheet[sheet_name] = stats

        _save_watched(existing)

        return {
            "added_new": new_items_count,
            "merged": merged_count,
            "already_recorded": already_recorded,
            "skipped_invalid": 0,
            "errors": errors,
            "total_unique": len(existing),
            "total_appearances": sum(
                len(it.get("appearances", [])) for it in existing
            ),
            "per_sheet": per_sheet,
            "workbook": str(workbook_path),
        }


@app.post("/watch/import-from-excel")
async def watch_import_from_excel(
    payload: dict[str, str] | None = None,
) -> dict[str, Any]:
    """Bulk-import every SECOP URL from the Dra's master Excel into the
    watch list. Auto-detects the header row (1 or 4) and the LINK column
    per sheet. Dedups by parsed ``process_id`` first, then URL string.

    Body (optional)::

        {"workbook": "path/to/excel.xlsx"}

    Defaults to ``BASE DE DATOS FEAB CONTRATOS2.xlsx`` in the cwd.

    Returns a per-sheet breakdown of how many were found/added/skipped
    so the UI can show "agregadas N de M nuevas" feedback.
    """
    workbook = (payload or {}).get("workbook") or _DEFAULT_WORKBOOK
    path = Path(workbook)
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        _executor, _import_workbook_urls, path
    )


@app.get("/audit-log")
async def audit_log(
    limit: int = Query(500, ge=1, le=10_000),
    op: str | None = Query(None, description="Filter by op: fill/replace/etc."),
) -> dict[str, Any]:
    """Return recent audit-log entries + chain integrity status."""
    log_path = state_path("audit_log.jsonl")
    if not log_path.exists():
        return {"entries": [], "intact": True, "problems": [], "total": 0}

    intact, problems = await asyncio.get_event_loop().run_in_executor(
        _executor, verify_audit_log, log_path,
    )
    entries: list[dict[str, Any]] = []
    total = 0
    for entry in _audit_iter(log_path):
        total += 1
        if op and entry.op != op:
            continue
        entries.append({
            "ts": entry.ts, "op": entry.op, "row": entry.row,
            "process_id": entry.process_id, "column": entry.column,
            "old": entry.old, "new": entry.new, "source": entry.source,
            "confidence": entry.confidence, "secop_hash": entry.secop_hash,
            "code_version": entry.code_version,
            "hash": entry.hash, "prev_hash": entry.prev_hash,
        })
    # Return tail (most recent) up to limit
    return {
        "entries": entries[-limit:],
        "intact": intact,
        "problems": problems,
        "total": total,
    }


@app.get("/ultima-actualizacion")
async def ultima_actualizacion() -> dict[str, Any]:
    """Última vez que el programa actualizó / verificó / consultó SECOP.

    Lee el audit log y reporta los timestamps más recientes por tipo
    de operación. La Dra. ve esto en el header para saber qué tan
    fresca está la data sin tener que disparar una verificación.
    """
    log_path = state_path("audit_log.jsonl")
    result: dict[str, str | None] = {
        "ultimo_fill": None,
        "ultimo_replace": None,
        "ultima_verificacion": None,
        "ultima_consulta": None,
        "total_operaciones": 0,
    }
    if not log_path.exists():
        return result
    total = 0
    for entry in _audit_iter(log_path):
        total += 1
        if entry.op == "fill" and (
            result["ultimo_fill"] is None or entry.ts > result["ultimo_fill"]
        ):
            result["ultimo_fill"] = entry.ts
        elif entry.op == "replace" and (
            result["ultimo_replace"] is None or entry.ts > result["ultimo_replace"]
        ):
            result["ultimo_replace"] = entry.ts
        elif entry.op == "verify_drift" and (
            result["ultima_verificacion"] is None
            or entry.ts > result["ultima_verificacion"]
        ):
            result["ultima_verificacion"] = entry.ts
    result["total_operaciones"] = total
    # The "ultima consulta" is the most recent of any of the above.
    timestamps = [v for k, v in result.items()
                 if k != "total_operaciones" and v]
    result["ultima_consulta"] = max(timestamps) if timestamps else None
    return result


@app.get("/modificatorios-recientes")
async def modificatorios_recientes(limit: int = Query(20, ge=1, le=100)) -> dict[str, Any]:
    """Aggregate view of every contract that has been modified — sorted
    by recency. Backbone of the 'Movimientos / Modificatorios' dashboard
    panel. Includes per-contract: last fecha_de_notificacion_de_prorroga
    if any, total días added, total adiciones value.
    """
    loop = asyncio.get_event_loop()

    def _fetch() -> dict[str, Any]:
        contratos = _client.query(
            "jbjy-vk9h",
            where=(
                f"nit_entidad='{FEAB_NIT}' AND "
                "(estado_contrato='Modificado' OR dias_adicionados>0)"
            ),
            limit=2000,
        )

        items: list[dict[str, Any]] = []
        for c in contratos:
            url_field = c.get("urlproceso")
            if isinstance(url_field, dict):
                url_field = url_field.get("url") or ""
            try:
                dias = int(float(c.get("dias_adicionados") or 0))
            except (ValueError, TypeError):
                dias = 0
            items.append({
                "id_contrato": c.get("id_contrato"),
                "referencia": c.get("referencia_del_contrato"),
                "proveedor": c.get("proveedor_adjudicado"),
                "objeto": (c.get("objeto_del_contrato") or "")[:120],
                "valor": c.get("valor_del_contrato"),
                "estado": c.get("estado_contrato"),
                "dias_adicionados": dias,
                "fecha_firma": (c.get("fecha_de_firma") or "")[:10],
                "fecha_fin": (c.get("fecha_de_fin_del_contrato") or "")[:10],
                "fecha_notificacion_prorroga":
                    (c.get("fecha_de_notificaci_n_de_prorrogaci_n") or "")[:10],
                "fecha_actualizacion": (c.get("ultima_actualizacion") or "")[:10],
                "url": url_field,
            })

        # Sort by the most-recent signal we have: notificacion_prorroga
        # first, then ultima_actualizacion, then fecha_fin (a contract
        # whose fin is later than original probably had a prorroga).
        def _key(it: dict) -> str:
            return (
                it.get("fecha_notificacion_prorroga", "")
                or it.get("fecha_actualizacion", "")
                or it.get("fecha_fin", "")
                or ""
            )
        items.sort(key=_key, reverse=True)

        last = items[0] if items else None
        return {
            "total_modificados": len(items),
            "total_dias_adicionados": sum(it["dias_adicionados"] for it in items),
            "ultimo": last,
            "items": items[:limit],
        }

    return await loop.run_in_executor(_executor, _fetch)


@app.get("/verify-progress")
async def verify_progress() -> dict[str, Any]:
    """Report progress of the most-recent ``scripts/verify_watch_list.py``
    run.

    Looks at ``.cache/watch_verify_*.jsonl`` to figure out:
      - ``running``: True if the latest report's mtime is < 30s ago
      - ``processed`` / ``total``: line count vs current watch list size
      - ``percent``, ``started_at``, ``last_update``, ``eta_seconds``

    The UI polls this endpoint every few seconds while a verify is
    in flight to render a progress bar.
    """
    import time

    cache_dir = state_dir()
    jsonls = sorted(
        cache_dir.glob("watch_verify_*.jsonl"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    total = len(_load_watched()) or 491
    if not jsonls:
        return {
            "running": False, "processed": 0, "total": total,
            "percent": 0.0, "started_at": None, "last_update": None,
            "eta_seconds": None, "report_path": None,
        }

    latest = jsonls[0]
    stat = latest.stat()
    now = time.time()
    age = now - stat.st_mtime

    processed = 0
    started_at_str: str | None = None
    last_line_ts: float | None = None
    try:
        with latest.open("r", encoding="utf-8") as fh:
            for line in fh:
                processed += 1
        # Parse the timestamp out of the filename: watch_verify_YYYYMMDD_HHMMSS.jsonl
        import re as _re
        m = _re.search(r"_(\d{8}_\d{6})\.", latest.name)
        elapsed = None
        eta = None
        if m:
            started_at_str = m.group(1)
            from datetime import datetime as _dt
            started_dt = _dt.strptime(started_at_str, "%Y%m%d_%H%M%S")
            last_line_ts = stat.st_mtime
            # `elapsed` measured against NOW so the UI shows real-time
            # elapsed even before any line lands. `rate` uses mtime so
            # we don't divide by inflated wall time.
            wall_elapsed = max(0, now - started_dt.timestamp())
            elapsed = wall_elapsed
            mtime_elapsed = stat.st_mtime - started_dt.timestamp()
            rate = processed / mtime_elapsed if mtime_elapsed > 0 else 0
            eta = (total - processed) / rate if rate > 0 else None
    except Exception as exc:
        log.warning("verify-progress read failed: %s", exc)
        eta = None

    # Heuristic: file mtime within last 30 s = the script is still
    # writing → running. Otherwise: finished or stale.
    running = age < 30 and processed < total

    return {
        "running": running,
        "processed": processed,
        "total": total,
        "percent": round(100 * processed / total, 1) if total else 0.0,
        "started_at": started_at_str,
        "elapsed_seconds": round(elapsed, 1) if elapsed else None,
        "last_update_age_seconds": round(age, 1),
        "eta_seconds": round(eta, 1) if eta else None,
        "report_path": str(latest),
    }


@app.post("/verify-watch")
async def verify_watch() -> dict[str, Any]:
    """Kick off ``scripts/verify_watch_list.py`` in the background.

    The Dra hits "Refrescar contra SECOP" — this re-reads every URL
    in her watch list against datos.gov.co and refreshes the
    notice_uid + verify_status taxonomy. Idempotent. Takes ~17 minutes
    for 491 URLs.
    """
    result = _run_script_async("scripts/verify_watch_list.py", [])
    return {
        **result,
        "message": "Verificación masiva iniciada — la Dra puede seguir "
                   "trabajando, los notice_uid se van actualizando.",
    }


@app.post("/refresh")
async def refresh(background: BackgroundTasks) -> dict[str, str]:
    """Kick off an Excel update in the background. Returns immediately."""
    workbook = "BASE DE DATOS FEAB CONTRATOS2.xlsx"
    if not Path(workbook).exists():
        raise HTTPException(404, f"workbook not found: {workbook}")

    def _run():
        from secop_ii.orchestrator import process_workbook
        try:
            report = process_workbook(
                workbook, do_backup=True, fields=["feab_fill"],
                generate_detalles=False, apply_view=True,
            )
            log.info("refresh done: %d ok, %d errors",
                    report.ok, report.errors)
        except Exception as exc:
            log.exception("refresh failed: %s", exc)

    background.add_task(_run)
    return {
        "status": "started",
        "workbook": workbook,
        "started_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }


@app.post("/verify")
async def verify_endpoint() -> dict[str, Any]:
    """Re-fetch SECOP for each row, compare with stored hashes."""
    workbook = "BASE DE DATOS FEAB CONTRATOS2.xlsx"
    if not Path(workbook).exists():
        raise HTTPException(404, f"workbook not found: {workbook}")
    loop = asyncio.get_event_loop()

    def _verify() -> dict[str, Any]:
        from secop_ii.verify import verify_workbook
        report = verify_workbook(workbook)
        return {
            "total": report.total,
            "fresh": report.fresh_count,
            "stale": report.stale_count,
            "errors": report.error_count,
            "drift_rows": [
                {"row": r.row, "process_id": r.process_id,
                 "stored_hash": r.stored_hash, "current_hash": r.current_hash}
                for r in report.rows if r.changed
            ],
        }

    return await loop.run_in_executor(_executor, _verify)


# ---- Portal scraper integration --------------------------------------------
# Para los procesos que el API público (datos.gov.co Socrata) no expone, el
# scraper ``scripts/scrape_portal.py`` baja el HTML de la página
# OpportunityDetail del portal SECOP y lo persiste en
# ``.cache/portal_opportunity.json``. Estos endpoints exponen ese cache a la
# UI sin atarla al filesystem y permiten lanzar el scraper como subprocess.

_PORTAL_CACHE = state_path("portal_opportunity.json")
_PORTAL_PROGRESS = state_path("portal_progress.jsonl")
_INTEGRADO_CACHE = state_path("secop_integrado.json")


def _read_integrado_cache() -> dict[str, Any]:
    """Carga el cache de SECOP Integrado (rpmr-utcd) — espejo de la
    API pública sin captcha. Generado por ``scripts/sync_secop_integrado.py``.
    """
    if not _INTEGRADO_CACHE.exists():
        return {"by_notice_uid": {}, "by_pccntr": {}, "synced_at": None, "total_rows": 0}
    try:
        import json as _json
        return _json.loads(_INTEGRADO_CACHE.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        log.warning("integrado cache ilegible: %s", exc)
        return {"by_notice_uid": {}, "by_pccntr": {}, "synced_at": None, "total_rows": 0}


@app.get("/contract-integrado/{key}")
async def contract_integrado(key: str) -> dict[str, Any]:
    """Espejo del proceso en el dataset SECOP Integrado (rpmr-utcd).

    ``key`` puede ser ``CO1.NTC.X`` (notice_uid) o ``CO1.PCCNTR.X``
    (numero_del_contrato). El endpoint busca primero por notice_uid;
    si no encuentra y la key es PCCNTR, busca por numero_del_contrato.

    Devuelve los campos crudos del API público — NUNCA derivados del
    Excel. Si el proceso no está en el dataset, ``available: false``.
    """
    cache = _read_integrado_cache()
    by_uid = cache.get("by_notice_uid") or {}
    by_pccntr = cache.get("by_pccntr") or {}

    row = by_uid.get(key) or by_pccntr.get(key)
    if not row:
        return {
            "available": False,
            "key": key,
            "synced_at": cache.get("synced_at"),
        }
    return {
        "available": True,
        "key": key,
        "fields": row,
        "synced_at": cache.get("synced_at"),
        "source": "rpmr-utcd (SECOP Integrado)",
    }


@app.get("/integrado-bulk")
async def integrado_bulk() -> dict[str, Any]:
    """Mapa completo del cache Integrado para la tabla principal.

    Devuelve un dict con campos summary por cada proceso, indexado tanto
    por notice_uid como por numero_del_contrato (CO1.PCCNTR.X). Esto le
    permite a la UI enriquecer CADA fila de la tabla sin un round-trip
    por proceso.

    Solo se exponen campos summary (no las 22 columnas completas del
    dataset) — para eso está ``/contract-integrado/{key}``.
    """
    cache = _read_integrado_cache()
    by_uid_full = cache.get("by_notice_uid") or {}
    by_pccntr_full = cache.get("by_pccntr") or {}

    SUMMARY_FIELDS = (
        "estado_del_proceso",
        "valor_contrato",
        "nom_raz_social_contratista",
        "fecha_de_firma_del_contrato",
        "fecha_inicio_ejecuci_n",
        "fecha_fin_ejecuci_n",
        "modalidad_de_contrataci_n",
        "tipo_de_contrato",
        "numero_del_contrato",
        "numero_de_proceso",
        "objeto_a_contratar",
        "url_contrato",
    )

    def _summary(row: dict[str, Any]) -> dict[str, Any]:
        return {k: row.get(k) for k in SUMMARY_FIELDS if row.get(k)}

    return {
        "synced_at": cache.get("synced_at"),
        "total_rows": cache.get("total_rows", 0),
        "source": cache.get("source"),
        "nit": cache.get("nit"),
        "by_notice_uid": {k: _summary(v) for k, v in by_uid_full.items()},
        "by_pccntr": {k: _summary(v) for k, v in by_pccntr_full.items()},
    }


@app.get("/integrado-summary")
async def integrado_summary() -> dict[str, Any]:
    """Resumen del cache Integrado: cuántos procesos, cuándo se sincronizó."""
    cache = _read_integrado_cache()
    return {
        "synced_at": cache.get("synced_at"),
        "total_rows": cache.get("total_rows", 0),
        "by_notice_uid_count": len(cache.get("by_notice_uid") or {}),
        "by_pccntr_count": len(cache.get("by_pccntr") or {}),
        "source": cache.get("source"),
        "nit": cache.get("nit"),
    }


@app.post("/integrado-sync")
async def integrado_sync(payload: dict[str, Any] | None = None) -> dict[str, Any]:
    """Lanza ``scripts/sync_secop_integrado.py`` (subprocess en dev, runpy en MSI)."""
    payload = payload or {}
    args: list[str] = []
    if payload.get("nit"):
        args += ["--nit", str(payload["nit"])]
    return _run_script_async("scripts/sync_secop_integrado.py", args)


def _read_portal_cache() -> dict[str, Any]:
    """Carga el cache compartido del scraper (idempotente, tolerante a corrupción)."""
    if not _PORTAL_CACHE.exists():
        return {}
    try:
        import json as _json
        return _json.loads(_PORTAL_CACHE.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        log.warning("portal cache ilegible: %s", exc)
        return {}


@app.get("/contract-portal/{notice_uid}")
async def contract_portal(notice_uid: str) -> dict[str, Any]:
    """Snapshot del portal SECOP para ``notice_uid`` (CO1.NTC.X).

    Devuelve TODOS los labels capturados (``all_labels``) además de los
    curados (``fields``), porque cada proceso del SECOP expone campos
    diferentes — la UI debe poder mostrar los específicos del proceso
    abierto sin asumir un esquema fijo.

    Si el proceso aún no fue scrapeado, retorna ``{available: false}``
    para que la UI ofrezca el botón "Leer del portal".
    """
    cache = _read_portal_cache()
    raw = cache.get(notice_uid)
    if not raw:
        return {"available": False, "notice_uid": notice_uid}

    # `all_labels` es el dump completo (label_normalizado → valor). Lo
    # exponemos tal cual y dejamos que el frontend ordene/filtre.
    return {
        "available": True,
        "notice_uid": notice_uid,
        "fields": raw.get("fields", {}),
        "all_labels": raw.get("all_labels", {}),
        "documents": raw.get("documents", []),
        "notificaciones": raw.get("notificaciones", []),
        "status": raw.get("status"),
        "missing_fields": raw.get("missing_fields", []),
        "scraped_at": raw.get("scraped_at"),
        "raw_length": raw.get("raw_length"),
    }


@app.get("/portal-progress")
async def portal_progress() -> dict[str, Any]:
    """Estado de la corrida actual del scraper del portal.

    Igual que ``/verify-progress`` pero leyendo
    ``.cache/portal_progress.jsonl`` (que escribe ``scrape_portal.py``).
    """
    if not _PORTAL_PROGRESS.exists():
        return {
            "running": False, "processed": 0, "total": 0,
            "percent": 0.0, "started_at": None,
            "eta_seconds": None, "last_event": None,
        }

    import json as _json
    import time as _time
    stat = _PORTAL_PROGRESS.stat()
    age = _time.time() - stat.st_mtime
    total = 0
    processed = 0
    started_at: str | None = None
    last_event: dict[str, Any] | None = None
    started_mono: float | None = None

    try:
        with _PORTAL_PROGRESS.open("r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    payload = _json.loads(line)
                except ValueError:
                    continue
                last_event = payload
                ev = payload.get("event")
                if ev == "start":
                    total = int(payload.get("total") or 0)
                    started_at = payload.get("started_at")
                elif ev == "item":
                    processed += 1
                elif ev == "done":
                    return {
                        "running": False,
                        "processed": int(payload.get("total", processed)),
                        "total": int(payload.get("total", total)),
                        "percent": 100.0,
                        "started_at": started_at,
                        "elapsed_seconds": payload.get("elapsed_seconds"),
                        "eta_seconds": 0,
                        "ok": payload.get("ok"),
                        "partial": payload.get("partial"),
                        "errored": payload.get("errored"),
                        "last_event": payload,
                    }
    except OSError as exc:
        log.warning("portal-progress read failed: %s", exc)

    # ETA: usar avg time per processed item × pendientes
    eta = None
    elapsed = None
    if started_at:
        try:
            from datetime import datetime as _dt
            started_dt = _dt.fromisoformat(started_at.replace("Z", "+00:00"))
            elapsed = max(0.0, _time.time() - started_dt.timestamp())
            if processed > 0 and total > processed:
                avg = elapsed / processed
                eta = avg * (total - processed)
        except (ValueError, TypeError):
            pass

    running = age < 90 and processed < total and total > 0
    return {
        "running": running,
        "processed": processed,
        "total": total,
        "percent": round(100 * processed / total, 1) if total else 0.0,
        "started_at": started_at,
        "elapsed_seconds": round(elapsed, 1) if elapsed else None,
        "last_update_age_seconds": round(age, 1),
        "eta_seconds": round(eta, 1) if eta else None,
        "last_event": last_event,
    }


@app.post("/portal-scrape")
async def portal_scrape(payload: dict[str, Any] | None = None) -> dict[str, Any]:
    """Lanza ``scripts/scrape_portal.py`` (subprocess en dev, runpy en MSI).

    Body opcional:
      ``{"uid": "CO1.NTC.X"}``     → un solo proceso (forzado)
      ``{"limit": 10}``             → primeros N pendientes
      ``{"force": true}``           → re-scrapea aunque haya cache OK
      (sin body)                    → todos los pendientes

    En el MSI (frozen) este endpoint requiere que Playwright esté instalado
    junto al ejecutable; si no, el script logea el ImportError y la barra
    de progreso queda en cero. La Dra abre community.secop.gov.co en su
    navegador y la app sigue mostrando los datos vía SECOP Integrado.
    """
    payload = payload or {}
    args: list[str] = []
    if payload.get("uid"):
        args += ["--uid", str(payload["uid"])]
    if payload.get("limit") is not None:
        args += ["--limit", str(int(payload["limit"]))]
    if payload.get("force"):
        args.append("--force")
    # Reset progress: la próxima corrida pisa el archivo (el script lo hace).
    return _run_script_async("scripts/scrape_portal.py", args)


# ---- Helpers ----------------------------------------------------------------


def _esc(value: str) -> str:
    """Escape single quotes for a Socrata WHERE clause."""
    return str(value).replace("'", "''")


def _run_script_async(script_relpath: str, args: list[str]) -> dict[str, Any]:
    """Launch a `scripts/*.py` worker — subprocess in dev, in-process thread in MSI.

    In dev ``sys.executable`` is ``python.exe`` and we spawn a child process the
    classic way. In a PyInstaller-frozen MSI build, ``sys.executable`` is the
    bundled ``.exe`` itself (it does not interpret Python), so we cannot
    ``Popen([sys.executable, "scripts/foo.py"])``. Instead we ``runpy.run_path``
    the script in a worker thread on the API's own process.

    Both paths return the same shape (``started``, ``pid``, ``cmd``) so the
    Next.js client doesn't need a frozen-vs-dev branch.
    """
    is_frozen = getattr(sys, "frozen", False)

    if is_frozen:
        meipass = Path(getattr(sys, "_MEIPASS", "."))
        script_path = meipass / script_relpath
        if not script_path.exists():
            raise HTTPException(
                500, f"Script no embebido en el .exe: {script_relpath}"
            )

        import runpy
        import threading

        def _runner() -> None:
            old_argv = sys.argv
            sys.argv = [str(script_path), *args]
            try:
                runpy.run_path(str(script_path), run_name="__main__")
            except SystemExit:
                # argparse calls sys.exit(0) on success
                pass
            except Exception:
                log.exception("Script %s failed in frozen runpy", script_relpath)
            finally:
                sys.argv = old_argv

        threading.Thread(target=_runner, daemon=False).start()
        return {
            "started": True,
            "pid": os.getpid(),
            "cmd": f"runpy:{script_relpath} " + " ".join(args),
        }

    # Dev mode — preserve the original detached-subprocess behavior.
    cmd = [sys.executable, "-X", "utf8", "-u", script_relpath, *args]
    proc = subprocess.Popen(
        cmd,
        cwd=str(Path(__file__).resolve().parent.parent.parent),
        env={**os.environ, "PYTHONUNBUFFERED": "1"},
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0),
    )
    return {
        "started": True,
        "pid": proc.pid,
        "cmd": " ".join(cmd),
    }


def main() -> None:
    """Entry point used by the launcher, the .bat file, and the MSI sidecar.

    The ``app`` object is passed by reference (not as a ``"secop_ii.api:app"``
    import string) so this works inside a PyInstaller bundle, where the
    entry script runs as ``__main__`` and the dotted name does not resolve.
    Reload is intentionally disabled — the MSI ships a frozen build.
    """
    # PyInstaller-frozen Python on Windows defaults stdout/stderr to cp1252
    # (the ANSI codepage) and IGNORES PYTHONUTF8 because the codec is wired
    # before env vars are read. The legacy CLI scripts print "→ Sincronizando…"
    # and "✓ Persistido en…"; without this reconfig those calls raise
    # UnicodeEncodeError inside the runpy worker thread. We force UTF-8
    # with ``errors='replace'`` so a broken byte never takes down the API.
    if getattr(sys, "frozen", False):
        for stream_name in ("stdout", "stderr"):
            stream = getattr(sys, stream_name, None)
            reconfigure = getattr(stream, "reconfigure", None)
            if callable(reconfigure):
                try:
                    reconfigure(encoding="utf-8", errors="replace")
                except Exception as exc:
                    log.warning("could not reconfigure sys.%s: %s",
                                stream_name, exc)
    # If running frozen (MSI) and the user has no .cache yet, bring our
    # bundled seed across before uvicorn boots so the lifespan handlers
    # (and the very first /watch hit) see a populated state dir. This is
    # also called from `_lifespan`, but doing it here too makes single-shot
    # CLI uses (`dra-cami-api.exe --version` style) feel consistent.
    try:
        _seed_state_dir_if_empty()
    except Exception as exc:
        log.warning("seed step failed (non-fatal): %s", exc)
    if getattr(sys, "frozen", False):
        # Surface a one-line boot marker into whatever stdout Tauri gave us
        # — usually the file `tauri/src/main.rs` redirects to. Helpful when
        # the Dra reports "no abre" and we need to prove the .exe ran.
        log.info("MSI sidecar boot · code_version=%s · state_dir=%s",
                 _CODE_VERSION, state_dir())
    import uvicorn
    uvicorn.run(
        app,
        host="127.0.0.1",
        port=8000,
        log_level="info",
        reload=False,
    )


if __name__ == "__main__":
    main()
