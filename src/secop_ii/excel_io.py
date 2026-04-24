"""Read/write Excel workbooks with openpyxl, preserving formatting.

We intentionally avoid going through pandas for writing because pandas'
``to_excel`` rebuilds the workbook from scratch and loses cell styles,
formulas, column widths and conditional formats. openpyxl edits cells
in-place, so the user's styles survive.

Key entry points:

* :func:`backup_workbook` — timestamped copy before any mutation.
* :func:`load_workbook` — opens and returns the workbook + active sheet.
* :func:`detect_url_column` — finds the column that holds SECOP URLs by
  header name (case/accent-insensitive) with a fallback to row scanning.
* :func:`ensure_columns` — appends new columns to the right of the data
  if they don't already exist, returning a name → column index map.
* :func:`iter_rows` — yields ``(row_index, url_value)`` skipping blanks.
* :func:`write_row` — writes a name → value dict into a given row.
"""

from __future__ import annotations

import re
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook as _openpyxl_load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

URL_HEADER_CANDIDATES = (
    "url",
    "urlproceso",
    "url proceso",
    "url secop",
    "url del proceso",
    "link",
    "link proceso",
    "link secop",
    "enlace",
    "enlace secop",
    "secop",
    "secop ii",
)

_SECOP_URL_RE = re.compile(
    r"https?://[^\s\"']*secop\.gov\.co/[^\s\"']*", re.IGNORECASE
)


class ExcelStructureError(RuntimeError):
    """Raised when the workbook does not have the expected shape."""


def backup_workbook(path: Path | str) -> Path:
    """Copy ``path`` to ``path.backup_<timestamp>.xlsx`` and return the copy.

    Raises:
        FileNotFoundError: ``path`` does not exist.
    """
    src = Path(path)
    if not src.is_file():
        raise FileNotFoundError(src)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dst = src.with_name(f"{src.stem}.backup_{stamp}{src.suffix}")
    shutil.copy2(src, dst)
    return dst


def load_workbook(
    path: Path | str, sheet_name: str | None = None
) -> tuple[Workbook, Worksheet]:
    """Open ``path`` and return ``(workbook, sheet)``.

    If ``sheet_name`` is ``None`` the active sheet is used.
    """
    wb = _openpyxl_load_workbook(str(path), data_only=False, keep_vba=False)
    ws = wb[sheet_name] if sheet_name else wb.active
    if ws is None:
        raise ExcelStructureError(f"No se encontró la hoja {sheet_name!r}")
    return wb, ws


def read_headers(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    """Return ``{header_name: column_index}`` for the given row.

    Empty headers are skipped. Column indices are 1-based (openpyxl).
    """
    result: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        value = cell.value
        if value is None or str(value).strip() == "":
            continue
        result[str(value).strip()] = col_idx
    return result


def detect_url_column(
    ws: Worksheet,
    header_row: int = 1,
    preferred: str | None = None,
) -> int:
    """Return the 1-based column index that holds SECOP URLs.

    Strategy:
        1. If ``preferred`` is given and matches a header, use it.
        2. Otherwise, scan headers for known candidates (``URL``, ``Link``,
           ``Enlace``, …) with case/accent-insensitive comparison.
        3. Fallback: scan the first 20 data rows and pick the column whose
           cells most often match a SECOP URL regex.

    Raises:
        ExcelStructureError: No column could be identified.
    """
    headers = read_headers(ws, header_row)

    if preferred:
        for name, idx in headers.items():
            if _norm(name) == _norm(preferred):
                return idx
        raise ExcelStructureError(
            f"La columna preferida {preferred!r} no existe. "
            f"Encabezados encontrados: {list(headers)}"
        )

    normalized_headers = {_norm(name): idx for name, idx in headers.items()}
    for candidate in URL_HEADER_CANDIDATES:
        idx = normalized_headers.get(_norm(candidate))
        if idx:
            return idx

    # Fallback — count SECOP URLs per column in the first 20 rows.
    best_col = None
    best_hits = 0
    for col_idx in range(1, ws.max_column + 1):
        hits = 0
        for r in range(header_row + 1, min(header_row + 21, ws.max_row + 1)):
            cell = ws.cell(row=r, column=col_idx).value
            if cell and _SECOP_URL_RE.search(str(cell)):
                hits += 1
        if hits > best_hits:
            best_hits = hits
            best_col = col_idx

    if best_col and best_hits > 0:
        return best_col

    raise ExcelStructureError(
        "No pude identificar una columna con URLs del SECOP. "
        "Asegúrate de que tu Excel tenga una columna llamada 'URL', 'Link', "
        "'Enlace' o similar con los enlaces a los procesos."
    )


def ensure_columns(
    ws: Worksheet,
    column_names: list[str] | tuple[str, ...],
    header_row: int = 1,
) -> dict[str, int]:
    """Guarantee that every ``column_names`` is present as a header.

    Existing columns keep their position; missing ones are appended to the
    right of the current data area. Returns ``{name: column_index}``.
    """
    existing = read_headers(ws, header_row)
    result: dict[str, int] = {}
    next_col = max(existing.values(), default=0) + 1
    for name in column_names:
        if name in existing:
            result[name] = existing[name]
            continue
        ws.cell(row=header_row, column=next_col, value=name)
        result[name] = next_col
        next_col += 1
    return result


def iter_rows(
    ws: Worksheet, url_col: int, header_row: int = 1
) -> Iterator[tuple[int, str]]:
    """Yield ``(row_index, url_value)`` for each non-blank data row."""
    for r in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(row=r, column=url_col).value
        if raw is None:
            continue
        text = str(raw).strip()
        if not text:
            continue
        yield r, text


def write_row(
    ws: Worksheet,
    row: int,
    values: dict[str, Any],
    column_map: dict[str, int],
) -> None:
    """Write ``values`` into ``row`` using ``column_map`` to resolve columns."""
    for name, value in values.items():
        col = column_map.get(name)
        if col is None:
            continue
        ws.cell(row=row, column=col, value=_normalize_cell_value(value))


def save_workbook(wb: Workbook, path: Path | str) -> None:
    wb.save(str(path))


def preview_as_dicts(
    ws: Worksheet, header_row: int = 1, limit: int | None = None
) -> list[dict[str, Any]]:
    """Return rows as plain dicts keyed by header name.

    Used by the Streamlit UI to render the CRM table.
    """
    headers = read_headers(ws, header_row)
    if not headers:
        return []
    inverse = {idx: name for name, idx in headers.items()}
    rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_dict: dict[str, Any] = {}
        has_any = False
        for col_idx, name in inverse.items():
            value = ws.cell(row=r, column=col_idx).value
            if value is not None and str(value).strip() != "":
                has_any = True
            row_dict[name] = value
        if has_any:
            row_dict["__row__"] = r
            rows.append(row_dict)
            if limit is not None and len(rows) >= limit:
                break
    return rows


def _norm(text: str) -> str:
    """Case + accent insensitive comparison helper."""
    stripped = unicodedata.normalize("NFKD", text)
    ascii_only = "".join(c for c in stripped if not unicodedata.combining(c))
    return ascii_only.strip().lower()


def _normalize_cell_value(value: Any) -> Any:
    """Return a value openpyxl can write without complaining."""
    if value is None:
        return None
    if isinstance(value, (int, float, str, bool)):
        return value
    if isinstance(value, (list, tuple)):
        return "; ".join(str(v) for v in value)
    return str(value)


__all__ = [
    "ExcelStructureError",
    "backup_workbook",
    "detect_url_column",
    "ensure_columns",
    "iter_rows",
    "load_workbook",
    "preview_as_dicts",
    "read_headers",
    "save_workbook",
    "write_row",
    "get_column_letter",
]
