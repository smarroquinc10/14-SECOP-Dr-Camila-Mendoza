"""Cross-check the Dra.'s OBSERVACIONES vs API vs Docs evidence per row.

Goal: detect every fila where the three signals disagree so the user can
review and decide whether (a) my classifier missed something, (b) the
SECOP II API lags behind the portal, or (c) the manual note was
aspirational/stale. No more false positives or false negatives.

Each row gets a verdict label from a small enum so the audit report can
be filtered to just the cases that need human eyes.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Verdict enum — what the three signals collectively say about this row.
# ---------------------------------------------------------------------------
# API + Docs (Socrata) son AUTORITATIVOS — vienen directo de SECOP II.
# La nota OBSERVACIONES de la Dra. es referencia humana, NO autoridad. El
# único caso que verdaderamente necesita revisión humana es:
# nota_dice_si pero ni API ni docs lo confirman (posible falso negativo).
# Cuando API/docs dicen sí, eso ES la verdad — la nota puede estar atrás.
VERDICT_AGREE_NO = "concuerdan_sin_modif"  # API=No, docs=No, nota=No
VERDICT_AGREE_SI = "concuerdan_con_modif"  # API+docs+nota concuerdan en Sí
VERDICT_API_LAG = "api_atrasado_docs_y_nota_si"  # docs+nota Sí, API normalizada No (lag)
VERDICT_NOTA_FALTA = "secop_dice_si_nota_no"  # API o docs Sí, nota No (la Dra. olvidó actualizar)
VERDICT_NOTE_ONLY = "nota_dice_si_secop_no"  # ⚠ ÚNICO caso que requiere revisión humana
VERDICT_DOCS_LIDERA = "docs_dice_si_api_no"  # docs Sí pero API formal No — probable lag
VERDICT_NA = "no_aplicable"  # url_invalida o sin proceso encontrado


@dataclass
class RowAudit:
    fila: int
    process_id: str | None
    objeto: str
    note_text: str
    note_says_modif: bool
    note_says_no_leg: bool
    api_says_modif: bool
    api_count: int
    docs_says_modif: bool
    docs_mod_count: int
    docs_leg_count: int
    docs_mod_list: str
    docs_leg_list: str
    verdict: str
    needs_review: bool

    def as_md_row(self) -> str:
        return (
            f"| {self.fila} | {self.process_id or '?'} | "
            f"{_short(self.objeto, 40)} | "
            f"{'Sí' if self.note_says_modif else 'No'}"
            f"{' (NO LEG)' if self.note_says_no_leg else ''} | "
            f"{'Sí' if self.api_says_modif else 'No'} ({self.api_count}) | "
            f"{self.docs_mod_count} mods + {self.docs_leg_count} legal | "
            f"**{self.verdict}**{' ⚠️' if self.needs_review else ''} |"
        )


def audit_workbook(path: Path | str) -> list[RowAudit]:
    """Load the enriched Excel and emit a verdict per fila."""
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    H = {h: i + 1 for i, h in enumerate(headers) if h}

    audits: list[RowAudit] = []
    for row_idx in range(2, ws.max_row + 1):
        get = lambda name: ws.cell(row=row_idx, column=H[name]).value if name in H else None

        process_id = get("ID identificado")
        if not process_id:
            continue  # blank row

        # Skip rows the program couldn't process at all.
        estado = (get("Estado actualización") or "").strip()
        if estado in ("url_invalida",):
            audits.append(_na_audit(row_idx, process_id, get))
            continue

        note_text = str(get("72. OBSERVACIONES") or "").strip()
        note_says_modif = (get("Modificatorio en OBS") or "") == "Sí"
        note_says_no_leg = (get("NO LEG") or "") == "Sí"

        api_modif_text = (get("¿Hubo modificatorio?") or "").strip()
        api_says_modif = api_modif_text == "Sí"
        api_count = _safe_int(get("# modificatorios"))

        docs_mod_count = _safe_int(get("Docs: # Modificatorios"))
        docs_leg_count = _safe_int(get("Docs: # Legalizaciones"))
        docs_says_modif = docs_mod_count > 0 or docs_leg_count > 0

        verdict = _verdict(note_says_modif, api_says_modif, docs_says_modif)
        # Solo NOTE_ONLY requiere ojo humano — los demás son hechos verificables.
        needs_review = verdict == VERDICT_NOTE_ONLY

        audits.append(
            RowAudit(
                fila=row_idx,
                process_id=process_id,
                objeto=str(get("Objeto en SECOP") or ""),
                note_text=note_text,
                note_says_modif=note_says_modif,
                note_says_no_leg=note_says_no_leg,
                api_says_modif=api_says_modif,
                api_count=api_count,
                docs_says_modif=docs_says_modif,
                docs_mod_count=docs_mod_count,
                docs_leg_count=docs_leg_count,
                docs_mod_list=str(get("Docs: Lista modificatorios") or ""),
                docs_leg_list=str(get("Docs: Lista legalizaciones") or ""),
                verdict=verdict,
                needs_review=needs_review,
            )
        )
    return audits


def _na_audit(row_idx: int, process_id: str, get) -> RowAudit:
    return RowAudit(
        fila=row_idx,
        process_id=process_id,
        objeto=str(get("Objeto en SECOP") or ""),
        note_text=str(get("72. OBSERVACIONES") or ""),
        note_says_modif=False,
        note_says_no_leg=False,
        api_says_modif=False,
        api_count=0,
        docs_says_modif=False,
        docs_mod_count=0,
        docs_leg_count=0,
        docs_mod_list="",
        docs_leg_list="",
        verdict=VERDICT_NA,
        needs_review=False,
    )


def _verdict(note: bool, api: bool, docs: bool) -> str:
    """Combine the three signals into one verdict label.

    SECOP II (API + docs) is authoritative; the OBSERVACIONES note is
    reference-only. So if SECOP says yes, the truth is yes regardless of
    what the note says — at most we flag "nota desactualizada".
    """
    secop_says = api or docs  # any positive signal from SECOP

    if not note and not secop_says:
        return VERDICT_AGREE_NO
    if note and api and docs:
        return VERDICT_AGREE_SI
    if note and not api and docs:
        # The fila 2 case: nota dice modificatorio, API normalizada laguea,
        # los PDFs publicados lo confirman.
        return VERDICT_API_LAG
    if note and api and not docs:
        # Nota + API agree; docs faltante (entidad no subió PDFs aún).
        return VERDICT_AGREE_SI
    if not note and secop_says:
        # SECOP autoritativo dice sí; la Dra. simplemente no lo registró.
        return VERDICT_NOTA_FALTA
    if note and not secop_says:
        # ⚠ Único caso que requiere revisión: nota dice algo que SECOP no
        # confirma. Posible: contrato fuera de SECOP, o nota errónea, o
        # mi clasificador falló al detectar un PDF.
        return VERDICT_NOTE_ONLY
    if not note and not api and docs:
        # Docs lidera, nota+API no. Casi siempre = API normalizada lagueando.
        return VERDICT_DOCS_LIDERA
    return VERDICT_AGREE_NO


def render_markdown(audits: list[RowAudit]) -> str:
    out: list[str] = []
    out.append("# Reporte de auditoría — Excel FEAB vs SECOP II\n")
    out.append(f"**Total filas analizadas:** {len(audits)}\n")

    # Summary by verdict
    by_verdict: dict[str, list[RowAudit]] = {}
    for a in audits:
        by_verdict.setdefault(a.verdict, []).append(a)

    out.append("## Resumen por veredicto\n")
    out.append("| Veredicto | # Filas | ¿Necesita revisión? |")
    out.append("|---|---|---|")
    for v, items in sorted(by_verdict.items(), key=lambda kv: -len(kv[1])):
        review = "✅ no" if not items[0].needs_review else "⚠️ **sí**"
        out.append(f"| `{v}` | {len(items)} | {review} |")
    out.append("")

    # Detail tables grouped by verdict
    for v, items in sorted(by_verdict.items(), key=lambda kv: -len(kv[1])):
        if not items:
            continue
        out.append(f"\n## {v} ({len(items)} filas)\n")
        out.append("| Fila | Proceso | Objeto | Tu nota | API | Docs | Veredicto |")
        out.append("|---|---|---|---|---|---|---|")
        for a in items[:30]:
            out.append(a.as_md_row())
        if len(items) > 30:
            out.append(f"| ... | ... | ... | ... | ... | ... | (+{len(items)-30} más) |")

    # Detailed evidence for every needs-review row
    flagged = [a for a in audits if a.needs_review]
    if flagged:
        out.append(f"\n---\n\n## 🔍 Casos que necesitan revisión humana ({len(flagged)})\n")
        for a in flagged:
            out.append(f"\n### Fila {a.fila} — {a.process_id} — `{a.verdict}`\n")
            out.append(f"**Objeto:** {_short(a.objeto, 200)}\n")
            out.append(f"**Tu OBSERVACIONES (intacta):** {_short(a.note_text, 350)}\n")
            out.append(
                f"- API `¿Hubo modificatorio?`: **{'Sí' if a.api_says_modif else 'No'}** "
                f"({a.api_count} modificatorios formales)"
            )
            out.append(
                f"- Docs SECOP: **{a.docs_mod_count}** modificatorios + "
                f"**{a.docs_leg_count}** legalizaciones"
            )
            if a.docs_mod_list:
                out.append(f"  - Mods: {_short(a.docs_mod_list, 250)}")
            if a.docs_leg_list:
                out.append(f"  - Legalizaciones: {_short(a.docs_leg_list, 250)}")
            out.append("")

    return "\n".join(out)


def _short(text: str, n: int) -> str:
    text = (text or "").replace("\n", " | ").strip()
    return text if len(text) <= n else text[: n - 1] + "…"


def _safe_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value).replace(",", "")))
    except (ValueError, TypeError):
        return 0


__all__ = ["RowAudit", "audit_workbook", "render_markdown"]
