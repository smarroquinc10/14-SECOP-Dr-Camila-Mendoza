"""Live verification: re-fetch SECOP and compare against stored hashes.

The Excel stores a SHA-256 fingerprint of the SECOP payload per row in
"FEAB: Hash SECOP (SHA-256)". Re-running the verifier:

* Re-fetches the live SECOP data for each row.
* Computes the current fingerprint.
* Compares with the stored one.
* Reports rows where the hash changed (= SECOP itself was modified
  since last update; the cells in the workbook may now be stale).

This is the "trust but verify" tool: the Dra. (or an auditor) can ask
"is what I have in the Excel STILL aligned with what SECOP says today?"
and get a yes/no per row, without re-running the full update.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from secop_ii.feab_columns import source_fingerprint
from secop_ii.secop_client import SecopClient
from secop_ii.url_parser import InvalidSecopUrlError, parse_secop_url

log = logging.getLogger(__name__)


@dataclass
class RowVerification:
    row: int
    process_id: str | None
    stored_hash: str | None
    current_hash: str | None
    changed: bool
    error: str | None = None


@dataclass
class VerificationReport:
    rows: list[RowVerification] = field(default_factory=list)

    @property
    def total(self) -> int:
        return len(self.rows)

    @property
    def stale_count(self) -> int:
        return sum(1 for r in self.rows if r.changed)

    @property
    def fresh_count(self) -> int:
        return sum(1 for r in self.rows if not r.changed and not r.error)

    @property
    def error_count(self) -> int:
        return sum(1 for r in self.rows if r.error)


def verify_workbook(
    path: Path | str,
    *,
    sheet_name: str | None = None,
    url_column_header: str = "LINK",
    hash_column_header: str = "FEAB: Hash SECOP (SHA-256)",
    progress: callable | None = None,
) -> VerificationReport:
    """Re-fetch SECOP for each row, compare hash with the one in the cell.

    Returns a report with stale/fresh/error counts.
    """
    wb = load_workbook(str(path), data_only=True, keep_vba=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = {c.value: i for i, c in enumerate(ws[1], start=1) if c.value}
    url_col = headers.get(url_column_header)
    hash_col = headers.get(hash_column_header)
    if not url_col:
        raise RuntimeError(f"Header '{url_column_header}' not found")
    if not hash_col:
        raise RuntimeError(
            f"Header '{hash_column_header}' not found — "
            "did you run an update first?"
        )

    client = SecopClient()
    report = VerificationReport()

    rows = [
        (r, ws.cell(row=r, column=url_col).value, ws.cell(row=r, column=hash_col).value)
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=url_col).value
    ]
    total = len(rows)

    for idx, (r, url, stored_hash) in enumerate(rows, start=1):
        try:
            ref = parse_secop_url(str(url))
            proceso = client.get_proceso(ref.process_id, url=ref.source_url)
            notice = client.resolve_notice_uid(ref.process_id, url=ref.source_url)
            portfolio = (proceso or {}).get("id_del_portafolio") or ""
            contratos = client.get_contratos(
                portfolio_id=str(portfolio), notice_uid=notice,
            )
            current = source_fingerprint(
                proceso=proceso, contratos=contratos, notice_uid=notice,
            )
            row_v = RowVerification(
                row=r,
                process_id=ref.process_id,
                stored_hash=str(stored_hash) if stored_hash else None,
                current_hash=current,
                changed=(stored_hash and str(stored_hash) != current),
            )
        except InvalidSecopUrlError as exc:
            row_v = RowVerification(
                row=r, process_id=None,
                stored_hash=str(stored_hash) if stored_hash else None,
                current_hash=None, changed=False,
                error=f"URL inválida: {exc}",
            )
        except Exception as exc:  # pragma: no cover - safety net
            log.exception("Verify falló en fila %s", r)
            row_v = RowVerification(
                row=r, process_id=None,
                stored_hash=str(stored_hash) if stored_hash else None,
                current_hash=None, changed=False,
                error=str(exc),
            )
        report.rows.append(row_v)
        if progress:
            progress(idx, total, row_v)

    return report


def render_verification_markdown(report: VerificationReport) -> str:
    """Render a human-readable Markdown summary of the verification."""
    lines = [
        "# Reporte de verificación — Dra Cami Contractual",
        "",
        f"**Total de filas verificadas:** {report.total}",
        f"**Coinciden con SECOP (frescas):** {report.fresh_count}",
        f"**SECOP cambió desde el último update (stale):** {report.stale_count}",
        f"**Errores:** {report.error_count}",
        "",
    ]
    if report.stale_count:
        lines.append("## Filas con drift (re-ejecutar update para refrescar)")
        lines.append("")
        lines.append("| Fila | Process ID | Hash anterior | Hash actual |")
        lines.append("|------|-----------|--------------|-------------|")
        for r in report.rows:
            if r.changed:
                lines.append(
                    f"| {r.row} | `{r.process_id}` | "
                    f"`{(r.stored_hash or '—')[:16]}…` | "
                    f"`{(r.current_hash or '—')[:16]}…` |"
                )
        lines.append("")
    if report.error_count:
        lines.append("## Errores")
        lines.append("")
        for r in report.rows:
            if r.error:
                lines.append(f"- Fila {r.row}: {r.error}")
        lines.append("")
    if report.stale_count == 0 and report.error_count == 0:
        lines.append("## Resultado")
        lines.append("")
        lines.append("Todas las filas están alineadas con SECOP. "
                    "No hay drift detectado.")
    return "\n".join(lines)


__all__ = [
    "RowVerification",
    "VerificationReport",
    "verify_workbook",
    "render_verification_markdown",
]
