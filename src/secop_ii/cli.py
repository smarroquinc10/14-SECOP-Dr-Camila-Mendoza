"""Command-line entry point for the Secop-II tool.

This CLI is used mostly for debugging and to let power users run the
tool without the Streamlit UI. The end-user deliverable is the Streamlit
``.exe`` that calls into the same building blocks.
"""

from __future__ import annotations

import json
import logging
import os
import sys
from typing import Annotated

import typer
from rich.console import Console
from rich.table import Table

from secop_ii.config import (
    DATASET_ADICIONES,
    DATASET_CONTRATOS,
    DATASET_PROCESOS,
    FIELD_ADICION_CONTRATO,
    FIELD_CONTRATO_PROCESO,
    FIELD_PROCESO_ID,
)
from secop_ii.extractors import REGISTRY, get_extractor
from secop_ii.extractors.base import ProcessContext
from secop_ii.orchestrator import process_workbook
from secop_ii.secop_client import SecopClient
from secop_ii.url_parser import InvalidSecopUrlError, parse_secop_url


def _run_all_extractors(ctx: ProcessContext) -> dict:
    """Execute every registered extractor and merge their outputs."""
    combined: dict = {}
    errors: list[str] = []
    all_ok = True
    for name in REGISTRY:
        try:
            result = get_extractor(name).extract(ctx)
        except Exception as exc:  # pragma: no cover
            errors.append(f"{name}: {exc}")
            all_ok = False
            continue
        combined.update(result.values)
        if not result.ok:
            all_ok = False
            if result.error:
                errors.append(f"{name}: {result.error}")
    return {"ok": all_ok, "error": "; ".join(errors) or None, "values": combined}

app = typer.Typer(
    help="Herramientas CLI para consultar procesos del SECOP II.",
    no_args_is_help=True,
)
console = Console()


def _configure_logging(verbose: bool) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )


@app.command("parse-url")
def parse_url_cmd(
    url: Annotated[str, typer.Argument(help="URL pública del SECOP II.")],
) -> None:
    """Parse a SECOP II URL and print the identifier it encodes."""
    try:
        ref = parse_secop_url(url)
    except InvalidSecopUrlError as exc:
        console.print(f"[red]ERROR:[/] {exc}")
        raise typer.Exit(code=2)

    console.print(f"[green]ID:[/]         {ref.process_id}")
    console.print(f"[green]Tipo:[/]       {ref.kind}")
    console.print(f"[green]Normalizada:[/] {ref.normalized_url}")


@app.command("show-queries")
def show_queries_cmd(
    url: Annotated[str, typer.Argument(help="URL pública del SECOP II.")],
) -> None:
    """Print the Socrata query URLs that would be used for this process.

    Useful when the machine running the CLI cannot reach datos.gov.co:
    copy any of the printed URLs into a browser that can, inspect the
    JSON, and share it back.
    """
    try:
        ref = parse_secop_url(url)
    except InvalidSecopUrlError as exc:
        console.print(f"[red]ERROR:[/] {exc}")
        raise typer.Exit(code=2)

    client = SecopClient()
    pid = ref.process_id

    console.print(f"[bold]Identificador:[/] {pid} ({ref.kind})\n")
    table = Table(title="Consultas Socrata que haría el programa")
    table.add_column("Dataset", style="cyan")
    table.add_column("URL", style="white", overflow="fold")
    table.add_row(
        f"Procesos ({DATASET_PROCESOS})",
        client.build_query_url(
            DATASET_PROCESOS, where=f"{FIELD_PROCESO_ID}='{pid}'", limit=1
        ),
    )
    table.add_row(
        f"Contratos ({DATASET_CONTRATOS})",
        client.build_query_url(
            DATASET_CONTRATOS,
            where=f"{FIELD_CONTRATO_PROCESO}='{pid}'",
            limit=50,
        ),
    )
    table.add_row(
        f"Adiciones ({DATASET_ADICIONES})",
        client.build_query_url(
            DATASET_ADICIONES,
            where=f"{FIELD_ADICION_CONTRATO}='<id_contrato de la consulta anterior>'",
            limit=50,
        ),
    )
    console.print(table)


@app.command("check-url")
def check_url_cmd(
    url: Annotated[str, typer.Argument(help="URL pública del SECOP II.")],
    app_token: Annotated[
        str | None,
        typer.Option(
            "--app-token",
            envvar="SOCRATA_APP_TOKEN",
            help="Token opcional de datos.gov.co (más cuota y menos 429).",
        ),
    ] = None,
    as_json: Annotated[bool, typer.Option("--json", help="Salida JSON.")] = False,
    verbose: Annotated[bool, typer.Option("-v", "--verbose")] = False,
) -> None:
    """Query SECOP II live and report whether the process has modificatorios."""
    _configure_logging(verbose)
    try:
        ref = parse_secop_url(url)
    except InvalidSecopUrlError as exc:
        console.print(f"[red]ERROR:[/] {exc}")
        raise typer.Exit(code=2)

    client = SecopClient(app_token=app_token)
    ctx = ProcessContext(ref=ref, client=client)
    outcome = _run_all_extractors(ctx)

    payload = {
        "id": ref.process_id,
        "kind": ref.kind,
        "ok": outcome["ok"],
        "error": outcome["error"],
        "values": outcome["values"],
    }

    if as_json:
        console.print_json(data=payload)
        return

    console.print(f"\n[bold]Proceso:[/] {ref.process_id}  ({ref.kind})")
    if not outcome["ok"] and outcome["error"]:
        console.print(f"[yellow]{outcome['error']}[/]")
    for col, val in outcome["values"].items():
        console.print(f"  [cyan]{col}:[/] {val}")


@app.command("check-json")
def check_json_cmd(
    url: Annotated[str, typer.Argument(help="URL original del SECOP II.")],
    proceso_file: Annotated[
        str | None,
        typer.Option(
            "--proceso",
            help="Ruta a JSON con la respuesta de datos.gov.co para el proceso.",
        ),
    ] = None,
    contratos_file: Annotated[
        str | None,
        typer.Option("--contratos", help="Ruta a JSON con los contratos."),
    ] = None,
    adiciones_file: Annotated[
        str | None,
        typer.Option("--adiciones", help="Ruta a JSON con las adiciones."),
    ] = None,
) -> None:
    """Run the modificatorios extractor using canned JSON fixtures.

    Intended for demos/tests when the machine running the CLI cannot
    reach datos.gov.co. Pass the JSON responses you downloaded manually
    from the browser and this command produces the same output that
    ``check-url`` would have produced against the live API.
    """
    ref = parse_secop_url(url)

    proceso = _load_json(proceso_file)
    if isinstance(proceso, list):
        proceso = proceso[0] if proceso else None
    contratos = _load_json(contratos_file) or []
    adiciones_all = _load_json(adiciones_file) or []

    class _OfflineContext(ProcessContext):
        def proceso(self_inner):  # type: ignore[override]
            return proceso

        def contratos(self_inner):  # type: ignore[override]
            return contratos

        def adiciones_de(self_inner, id_contrato):  # type: ignore[override]
            return [
                a
                for a in adiciones_all
                if str(a.get(FIELD_ADICION_CONTRATO)) == str(id_contrato)
            ]

    ctx = _OfflineContext(ref=ref, client=SecopClient())
    outcome = _run_all_extractors(ctx)

    console.print(f"\n[bold]Proceso:[/] {ref.process_id} ({ref.kind})  [offline demo]")
    if not outcome["ok"] and outcome["error"]:
        console.print(f"[yellow]{outcome['error']}[/]")
    for col, val in outcome["values"].items():
        console.print(f"  [cyan]{col}:[/] {val}")


def _load_json(path: str | None):
    if not path:
        return None
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


@app.command("update-excel")
def update_excel_cmd(
    input_file: Annotated[str, typer.Argument(help="Ruta al archivo .xlsx a actualizar.")],
    url_column: Annotated[
        str | None,
        typer.Option(
            "--url-column",
            help="Encabezado exacto de la columna con las URLs (autodetecta si no se pasa).",
        ),
    ] = None,
    sheet: Annotated[
        str | None,
        typer.Option("--sheet", help="Nombre de la hoja (por defecto la activa)."),
    ] = None,
    header_row: Annotated[
        int, typer.Option("--header-row", help="Fila del encabezado (1-based).")
    ] = 1,
    app_token: Annotated[
        str | None,
        typer.Option(
            "--app-token",
            envvar="SOCRATA_APP_TOKEN",
            help="Token opcional de datos.gov.co.",
        ),
    ] = None,
    no_backup: Annotated[
        bool, typer.Option("--no-backup", help="Saltar la creación del backup.")
    ] = False,
    dry_run: Annotated[
        bool, typer.Option("--dry-run", help="No escribe cambios al archivo."),
    ] = False,
    no_portal: Annotated[
        bool,
        typer.Option(
            "--no-portal",
            help=(
                "Saltar el espejo profundo del portal SECOP II. Por defecto "
                "el programa abre Chrome visible y scrapea cada "
                "OpportunityDetail (pide 1 clic al captcha por sesión). "
                "Usa este flag si solo quieres datos del API datos.gov.co."
            ),
        ),
    ] = False,
    all_sheets: Annotated[
        bool,
        typer.Option(
            "--all-sheets",
            help=(
                "Itera por TODAS las hojas del libro (FEAB 2026, 2025, 2024, "
                "2023, 2022, 2018-2021). Auto-detecta header_row 4 para "
                "'FEAB 2018-2021' (formato gestión contractual con títulos "
                "en filas 1-3). Sin este flag solo procesa la hoja activa."
            ),
        ),
    ] = False,
    verbose: Annotated[bool, typer.Option("-v", "--verbose")] = False,
) -> None:
    """Actualiza un Excel con datos del SECOP II."""
    _configure_logging(verbose)

    def _progress(done: int, total: int, row) -> None:
        mark = "[OK]" if row.ok else "[X] "
        console.print(
            f"[dim]{done:>3}/{total}[/] {mark} fila {row.row}: "
            f"{row.process_id or row.url[:60]} -> {row.status}"
        )

    if all_sheets:
        from openpyxl import load_workbook as _lw
        wb = _lw(input_file, read_only=True)
        sheet_names = list(wb.sheetnames)
        wb.close()
        console.print(
            f"[bold]--all-sheets:[/] procesando {len(sheet_names)} hojas: "
            f"{', '.join(sheet_names)}"
        )

        # Backup once at the start, then disable per-sheet backups so we
        # don't generate 6 timestamped copies.
        do_backup_first = not no_backup
        all_errors = 0
        all_ok = 0
        for sn in sheet_names:
            # FEAB 2018-2021 (legacy format) keeps headers in row 4.
            sheet_header_row = 4 if sn.strip() == "FEAB 2018-2021" else header_row
            console.print(
                f"\n[bold cyan]>>> Hoja:[/] {sn} (header_row={sheet_header_row})"
            )
            try:
                report_i = process_workbook(
                    input_file,
                    url_column=url_column,
                    sheet_name=sn,
                    header_row=sheet_header_row,
                    app_token=app_token,
                    do_backup=do_backup_first,
                    progress=_progress,
                    dry_run=dry_run,
                    mirror_portal=not no_portal,
                )
                all_ok += report_i.ok
                all_errors += report_i.errors
            except Exception as exc:
                console.print(f"[red]Error en hoja {sn}: {exc}[/]")
                all_errors += 1
            do_backup_first = False  # only backup before the first sheet

        console.print(
            f"\n[bold]Resumen --all-sheets:[/] {all_ok} ok, {all_errors} errores"
        )
        if all_errors:
            raise typer.Exit(code=1)
        return

    report = process_workbook(
        input_file,
        url_column=url_column,
        sheet_name=sheet,
        header_row=header_row,
        app_token=app_token,
        do_backup=not no_backup,
        progress=_progress,
        dry_run=dry_run,
        mirror_portal=not no_portal,
    )

    console.print("\n[bold]Resumen:[/]")
    summary = report.as_dict()
    for key, value in summary.items():
        console.print(f"  [cyan]{key}:[/] {value}")
    if report.errors:
        raise typer.Exit(code=1)


@app.command("report")
def report_cmd(
    input_file: Annotated[str, typer.Argument(help="Excel ya enriquecido (después de update-excel).")],
    output: Annotated[
        str | None,
        typer.Option("--output", "-o", help="Ruta de salida; por defecto AUDITORIA_<fecha>.xlsx en el mismo dir."),
    ] = None,
    md_too: Annotated[
        bool, typer.Option("--md", help="Generar también el AUDITORIA.md clásico junto al .xlsx."),
    ] = False,
) -> None:
    """Generate a professional auditing Excel from the enriched workbook.

    The output is a 3-sheet workbook (Resumen, Detalle, Banderas Rojas)
    with conditional formatting, autofilters, frozen panes — designed to
    be handed to a supervisor or compliance officer without further edits.
    """
    from secop_ii.audit import audit_workbook, render_markdown
    from secop_ii.excel_pro import build_audit_workbook

    src = os.path.abspath(input_file)
    audits = audit_workbook(src)

    if output is None:
        from datetime import datetime as _dt
        stem = _dt.now().strftime("AUDITORIA_%Y-%m-%d_%H%M.xlsx")
        output = os.path.join(os.path.dirname(src) or ".", stem)

    info = build_audit_workbook(audits, output, excel_source=os.path.basename(src))
    console.print(f"[green]OK[/]  Reporte Excel generado: [bold]{output}[/]")
    console.print(
        f"      {info.total_rows} filas | {info.needs_review} requieren revisión humana"
    )

    if md_too:
        md_path = os.path.splitext(output)[0] + ".md"
        from pathlib import Path as _P
        _P(md_path).write_text(render_markdown(audits), encoding="utf-8")
        console.print(f"[green]OK[/]  Markdown:       {md_path}")


@app.command("inspect")
def inspect_cmd(
    process_id: Annotated[str, typer.Argument(help="Identificador (CO1.PPI.* / CO1.NTC.* / CO1.REQ.*) o URL.")],
    download: Annotated[
        bool,
        typer.Option(
            "--download/--no-download",
            help="Descargar y leer cada PDF publicado (cacheado en .cache/pdf).",
        ),
    ] = True,
    max_pdfs: Annotated[
        int, typer.Option("--max-pdfs", help="Máximo de PDFs a descargar/parsear."),
    ] = 8,
    note: Annotated[
        str | None,
        typer.Option(
            "--note",
            help="Texto OBSERVACIONES de la Dra. (opcional). Se compara fuzzy con los nombres de archivo.",
        ),
    ] = None,
    verbose: Annotated[bool, typer.Option("-v", "--verbose")] = False,
) -> None:
    """Deep-dive on a single process: API + docs + PDF peek + fuzzy match.

    This is the human-eyes mode for cases the audit flagged. It pulls
    every record from every relevant Socrata dataset, downloads the
    published PDFs (cached) and surfaces a compact report.
    """
    _configure_logging(verbose)
    from secop_ii.notice_resolver import NoticeResolver
    from secop_ii.pdf_reader import PdfReader
    from rapidfuzz import fuzz

    pid = process_id.strip()
    is_url = pid.startswith("http")

    client = SecopClient()
    if is_url:
        try:
            ref = parse_secop_url(pid)
            pid = ref.process_id
        except InvalidSecopUrlError as exc:
            console.print(f"[red]ERROR[/] {exc}")
            raise typer.Exit(code=2)

    # 1) Resolve PPI -> NTC if applicable
    ntc = None
    if pid.startswith("CO1.PPI."):
        nr = NoticeResolver()
        ntc = nr.resolve(f"https://community.secop.gov.co/Public/Tendering/ContractNoticePhases/View?PPI={pid}")
        console.print(f"[cyan]PPI[/] {pid}  ->  NTC {ntc or '(no resolvió)'}")

    # 2) Process row
    proc = client.get_proceso(
        pid,
        url=f"https://community.secop.gov.co/Public/Tendering/ContractNoticePhases/View?PPI={pid}"
        if pid.startswith("CO1.PPI.") else None,
    )
    if proc is None:
        console.print("[yellow]No se encontró proceso en Socrata.[/]")
    else:
        portfolio = proc.get("id_del_portafolio") or "?"
        console.print(f"[cyan]Proceso[/]  fase={proc.get('fase')}  adjudicado={proc.get('adjudicado')}  portafolio={portfolio}")
        console.print(f"           objeto: {(proc.get('descripci_n_del_procedimiento') or '')[:120]}")

    # 3) Contratos
    contratos = []
    if proc and proc.get("id_del_portafolio"):
        contratos = client.get_contratos(portfolio_id=str(proc["id_del_portafolio"]))
    if not contratos and ntc:
        contratos = client.get_contratos(notice_uid=ntc)
    console.print(f"[cyan]Contratos[/] encontrados: {len(contratos)}")
    for c in contratos[:5]:
        console.print(
            f"           id_ctr={c.get('id_contrato')}  estado={c.get('estado_contrato')}  "
            f"valor={c.get('valor_del_contrato')}"
        )

    # 4) Archivos publicados (Socrata)
    archivos = []
    if proc and proc.get("id_del_portafolio"):
        archivos = client.get_archivos(str(proc["id_del_portafolio"]))
    console.print(f"[cyan]Archivos[/] publicados: {len(archivos)}")

    table = Table(title="PDFs publicados")
    table.add_column("Fecha", style="dim", no_wrap=True)
    table.add_column("Nombre", style="white", overflow="fold")
    table.add_column("Engine", style="cyan", no_wrap=True)
    table.add_column("Mod KW", style="yellow", no_wrap=True)
    table.add_column("Match nota", style="magenta", no_wrap=True)

    reader = PdfReader() if download else None
    for arc in archivos[:max_pdfs]:
        nm = arc.get("nombre_archivo") or ""
        fc = (arc.get("fecha_carga") or "")[:10]
        url_field = arc.get("_url_normalized") or arc.get("url_descarga_documento") or ""
        if isinstance(url_field, dict):
            url_field = url_field.get("url", "")

        engine = "—"
        mod_hits = ""
        match_score = ""
        if reader and url_field:
            summary = reader.summarize(url_field)
            engine = summary.engine
            mod_hits = ",".join(summary.keywords_modif) if summary.keywords_modif else ""
            if note:
                # Fuzzy match nombre + preview vs OBSERVACIONES note
                best = max(
                    fuzz.partial_ratio(note, nm),
                    fuzz.partial_ratio(note, summary.text_preview[:300]) if summary.text_preview else 0,
                )
                match_score = f"{best}%"
        table.add_row(fc, nm[:80], engine, mod_hits[:40], match_score)
    console.print(table)

    # 5) Verdict on the spot
    docs_says = sum(
        1 for a in archivos
        if any(kw in (a.get("nombre_archivo") or "").upper() for kw in ("MODIFIC", "OTROSI", "PRORROGA", "ADICION"))
    )
    api_says = len(contratos) > 0  # very rough
    console.print()
    console.print("[bold]Resumen rápido:[/]")
    console.print(f"  proceso encontrado:      {'sí' if proc else 'no'}")
    console.print(f"  contratos en Socrata:    {len(contratos)}")
    console.print(f"  archivos en Socrata:     {len(archivos)}")
    console.print(f"  filename con keyword mod: {docs_says}")
    if not proc or (len(contratos) == 0 and len(archivos) == 0):
        console.print("[red bold]  -> SECOP no tiene evidencia. Si tu nota afirma modificatorio,[/red bold]")
        console.print("[red bold]     es un caso de revisión humana (archivo físico / otro NTC).[/red bold]")


@app.command("export")
def export_cmd(
    source_excel: Annotated[
        str,
        typer.Argument(help="Excel de origen (provee la lista de URLs SECOP)."),
    ],
    out: Annotated[
        str,
        typer.Option("--out", "-o", help="Excel de salida con la espejo SECOP limpio."),
    ] = "Dra_Cami_Contractual_export.xlsx",
    sheet_name: Annotated[
        str | None, typer.Option("--sheet", help="Hoja del origen."),
    ] = None,
    detalles: Annotated[
        bool, typer.Option("--detalles/--no-detalles", help="Generar HTML drill-down."),
    ] = True,
    verbose: Annotated[bool, typer.Option("-v", "--verbose")] = False,
) -> None:
    """Genera un Excel LIMPIO desde cero con SECOP como única fuente.

    Usa el archivo origen solo para extraer las URLs de los procesos.
    Crea un nuevo workbook donde cada celda viene de SECOP — sin data
    legacy. Útil para auditorías, entregables a leadership, o reset.
    """
    _configure_logging(verbose)
    from openpyxl import Workbook, load_workbook as _open
    from secop_ii.feab_columns import FEAB_COLUMNS_ORDERED
    from secop_ii.excel_io import detect_url_column
    from pathlib import Path
    import shutil

    src = Path(source_excel)
    dst = Path(out)
    if not src.exists():
        console.print(f"[red]ERROR[/] no existe {src}")
        raise typer.Exit(code=2)

    # Strategy: copy the source as starting point so column structure
    # and styles are preserved, then run the orchestrator on the copy
    # with the FEAB filler — but we need to wipe the her-77 columns
    # first so SECOP's values are pure (no merge with legacy).
    shutil.copy(src, dst)
    console.print(f"[green]OK[/]  Copia base: {dst}")

    wb = _open(str(dst))
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = {c.value: i for i, c in enumerate(ws[1], start=1) if c.value}
    # Wipe SECOP-derivable columns so the filler writes from scratch.
    # Internal columns stay (CDP, Orfeo, Abogado — those are FEAB-only).
    from secop_ii.feab_columns import INTERNAL_ONLY
    wiped = 0
    for col_name, col_idx in headers.items():
        if col_name in FEAB_COLUMNS_ORDERED and col_name not in INTERNAL_ONLY:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if cell.value not in (None, ""):
                    cell.value = None
                    wiped += 1
    wb.save(str(dst))
    console.print(f"[green]OK[/]  Limpiadas {wiped} celdas (SECOP-derivables) para pinta fresca")

    # Now run the orchestrator on the wiped copy. With no manual values
    # in those cells, every fill is a fresh SECOP write — pure mirror.
    console.print(f"[bold]Pintando[/] desde SECOP en vivo…")
    report = process_workbook(
        dst, do_backup=False, fields=["feab_fill"],
        generate_detalles=detalles, apply_view=True,
    )
    console.print(
        f"[bold green]Export listo[/] · {report.ok}/{report.total} filas OK · "
        f"errores: {report.errors}"
    )
    console.print(f"  Excel: {dst}")
    if detalles:
        console.print(f"  Detalles: {dst.parent / 'detalles'}/")


@app.command("verify")
def verify_cmd(
    excel_path: Annotated[str, typer.Argument(help="Ruta del Excel con hashes.")],
    out: Annotated[
        str | None,
        typer.Option("--out", help="Si se da, escribe el reporte en Markdown."),
    ] = None,
    verbose: Annotated[bool, typer.Option("-v", "--verbose")] = False,
) -> None:
    """Re-pega contra SECOP y compara hashes con los del Excel.

    Detecta filas donde SECOP cambió desde el último update — esas
    filas están "stale" y conviene re-correr update-excel para
    refrescar. Filas frescas = todavía espejo de SECOP.
    """
    _configure_logging(verbose)
    from secop_ii.verify import render_verification_markdown, verify_workbook

    console.print(f"[bold]Verificando[/] {excel_path} contra SECOP en vivo…")

    def _progress(idx: int, total: int, row_v) -> None:
        status = (
            "[red]drift[/]" if row_v.changed
            else "[yellow]error[/]" if row_v.error
            else "[green]fresh[/]"
        )
        console.print(f"  {idx}/{total} fila {row_v.row} {row_v.process_id} → {status}")

    report = verify_workbook(excel_path, progress=_progress)

    console.print()
    console.print("[bold]Resultado[/]")
    console.print(f"  Total:         {report.total}")
    console.print(f"  [green]Frescas:[/]      {report.fresh_count}")
    console.print(f"  [red]Con drift:[/]    {report.stale_count}")
    console.print(f"  [yellow]Errores:[/]      {report.error_count}")

    if out:
        from pathlib import Path
        Path(out).write_text(render_verification_markdown(report), encoding="utf-8")
        console.print(f"[green]OK[/]  Reporte: {out}")


@app.command("audit-log")
def audit_log_cmd(
    log_path: Annotated[
        str, typer.Argument(help="Ruta del audit log JSONL."),
    ] = ".cache/audit_log.jsonl",
    verify_chain: Annotated[
        bool,
        typer.Option("--verify/--no-verify", help="Verificar la integridad del hash-chain."),
    ] = True,
    out: Annotated[
        str | None, typer.Option("--out", help="Markdown con el resumen."),
    ] = None,
) -> None:
    """Auditoría inmutable: muestra y verifica el log hash-chained.

    Cada operación que escribe en el Excel deja una entrada en el log
    enlazada por SHA-256 con la anterior. Si alguien edita una entrada
    pasada, la cadena se rompe y este comando lo reporta.
    """
    from pathlib import Path
    from secop_ii.audit_log import render_audit_summary, verify_audit_log

    p = Path(log_path)
    if not p.exists():
        console.print(f"[yellow]Aviso:[/] no existe {log_path} — no hay log aún.")
        raise typer.Exit(code=0)

    summary = render_audit_summary(p)
    console.print(summary)

    if verify_chain:
        intact, problems = verify_audit_log(p)
        console.print()
        if intact:
            console.print("[green bold]Hash-chain íntegro[/] — ninguna alteración detectada.")
        else:
            console.print("[red bold]ALERTA: cadena rota[/]")
            for p_msg in problems:
                console.print(f"  [red]·[/] {p_msg}")

    if out:
        Path(out).write_text(summary, encoding="utf-8")
        console.print(f"[green]OK[/]  Resumen: {out}")


if __name__ == "__main__":  # pragma: no cover
    app()
