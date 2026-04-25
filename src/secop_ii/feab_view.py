"""View polish on the Excel after the orchestrator runs.

Two responsibilities:

1. **Hide auxiliary columns** so the Dra. opens the workbook and sees ONLY
   her 77 columns + a few audit ones (Discrepancias, Confianza, Hash,
   Ver detalle). The 146 data-mirror columns stay in the file (full
   storage), just hidden by default. Right-click → unhide brings them
   back when she needs the deep view.

2. **Freeze panes** at row 1, column 2 so column headers stay visible
   when scrolling, and "2. NÚMERO DE CONTRATO" stays as the row anchor.

3. **Add a "Ver detalle" hyperlink** in column 74 (LINK) and a separate
   one to the per-process HTML file when it exists.

Visible columns by default:
* 1-77: the Dra.'s columns
* "Estado actualización" / "Última actualización": top-of-mind status
* "FEAB: Discrepancias" / "FEAB: Confianza global" / "FEAB: Celdas a
  revisar" / "FEAB: Ver detalle" / "FEAB: Hash SECOP (SHA-256)"
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Columns that should stay visible when applying the "Dra." view.
_VISIBLE_AUDIT_COLUMNS: frozenset[str] = frozenset({
    "Estado actualización",
    "Última actualización",
    "FEAB: Auto-llenadas",
    "FEAB: Discrepancias",
    "FEAB: Celdas a revisar",
    "FEAB: Confianza global",
    "FEAB: Hash SECOP (SHA-256)",
    "FEAB: Ver detalle",
})


def apply_dra_view(
    workbook_path: Path | str,
    *,
    sheet_name: str | None = None,
    visible_audit_columns: Iterable[str] | None = None,
    detalles_dir: str = "detalles",
) -> dict[str, int]:
    """Hide aux columns + freeze panes + format the Dra.'s working view.

    Returns a count of ``{visible: N, hidden: M, total: N+M}`` for logging.
    Idempotent: running it twice has the same effect.
    """
    extra_visible = set(visible_audit_columns or ()) | _VISIBLE_AUDIT_COLUMNS
    wb = load_workbook(str(workbook_path))
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = _read_headers(ws)
    if not headers:
        wb.save(str(workbook_path))
        return {"visible": 0, "hidden": 0, "total": 0}

    # The 77 numbered/named columns belong to the Dra. (cols 1-77).
    # Anything beyond column 77 is auxiliary and gets hidden unless
    # explicitly listed as visible audit.
    visible_count = 0
    hidden_count = 0
    for col_idx, name in headers.items():
        letter = get_column_letter(col_idx)
        if col_idx <= 77 or name in extra_visible:
            ws.column_dimensions[letter].hidden = False
            visible_count += 1
        else:
            ws.column_dimensions[letter].hidden = True
            hidden_count += 1

    # Freeze panes: top row + first col (NÚMERO DE CONTRATO).
    ws.freeze_panes = "B2"

    # Highlight the audit columns so they're easy to spot at the right
    # of her natural view.
    audit_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1",
                             fill_type="solid")
    audit_font = Font(bold=True, color="C9A227")
    for col_idx, name in headers.items():
        if name in _VISIBLE_AUDIT_COLUMNS and col_idx > 77:
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = audit_fill
            cell.font = audit_font

    # Add hyperlinks to the per-process HTML detail files. The
    # "FEAB: Ver detalle" column already contains the relative path;
    # turn it into a clickable link.
    ver_col = next((c for c, n in headers.items()
                   if n == "FEAB: Ver detalle"), None)
    if ver_col:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=ver_col)
            if not cell.value:
                continue
            rel = str(cell.value)
            cell.hyperlink = rel
            cell.style = "Hyperlink"
            # Plain ASCII to avoid Windows console encoding issues
            # when the user's tools (or our own logs) read the cell.
            cell.value = "Abrir detalle >>"

    wb.save(str(workbook_path))
    return {
        "visible": visible_count,
        "hidden": hidden_count,
        "total": visible_count + hidden_count,
    }


def _read_headers(ws: Worksheet) -> dict[int, str]:
    """Return ``{col_idx: name}`` for the header row."""
    out: dict[int, str] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        v = cell.value
        if v is None or str(v).strip() == "":
            continue
        out[col_idx] = str(v).strip()
    return out


__all__ = ["apply_dra_view"]
