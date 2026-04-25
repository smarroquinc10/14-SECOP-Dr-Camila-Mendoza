"""Track what changed in the Excel between updates.

The Dra. is the head of contratación del FEAB — when she opens the CRM
she wants to know *at a glance* what has moved since she last looked.
That means a per-day snapshot of the key fields of every process, and a
diff when a new snapshot is computed.

Design:

* A snapshot is ``{process_id: {key_field: value}}`` for a small set of
  watched fields (state, phase, value, mod count, supplier, etc.).
* Snapshots are stored as one JSON file per day under
  ``.cache/snapshots/YYYY-MM-DD.json``. New snapshots overwrite the
  same-day file so multiple updates in one day collapse to "today".
* :func:`diff_snapshots` returns a :class:`Changelog` with added,
  removed and changed processes — each change lists which fields moved
  and from-to values.
* :func:`summarize_changelog` produces the one-line "N procesos nuevos,
  M cambios de estado, K modificatorios" string for the welcome card.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook as _load

# Fields we watch for changes. Picked because:
#   - "Fase en SECOP" / "Contrato: Estado" — legal state changes
#   - "Valor adjudicación" / "Contrato: Valor" / "Contrato: Valor pagado"
#     — money flow changes
#   - "¿Hubo modificatorio?" / "# modificatorios" — new modificatorios
#   - "Proveedor adjudicado" — a change here is a huge red flag
#   - "Docs: # Modificatorios" / "Docs: # Legalizaciones" — new published PDFs
_WATCHED_FIELDS = (
    "Fase en SECOP",
    "Contrato: Estado",
    "Proceso: Valor adjudicación",
    "Contrato: Valor",
    "Contrato: Valor pagado",
    "¿Hubo modificatorio?",
    "# modificatorios",
    "Contrato: Proveedor adjudicado",
    "Docs: # Modificatorios",
    "Docs: # Legalizaciones",
    "Pagos: # facturas",
    "Pagos: Total pagado",
)

_ID_COLUMN = "ID identificado"
_SNAPSHOT_DIR = Path(".cache") / "snapshots"


@dataclass
class FieldChange:
    field: str
    old: Any
    new: Any


@dataclass
class ProcessChange:
    process_id: str
    changes: list[FieldChange]


@dataclass
class Changelog:
    prev_date: str                        # date of the snapshot we compared against
    added: list[str] = field(default_factory=list)        # new process_ids
    removed: list[str] = field(default_factory=list)      # process_ids deleted
    changed: list[ProcessChange] = field(default_factory=list)  # field-level changes

    @property
    def total(self) -> int:
        return len(self.added) + len(self.removed) + len(self.changed)

    @property
    def is_empty(self) -> bool:
        return self.total == 0


# ---------------------------------------------------------------------------
# Snapshot I/O
# ---------------------------------------------------------------------------
def snapshot_from_excel(path: Path | str, header_row: int = 1) -> dict[str, dict[str, Any]]:
    """Extract ``{process_id: {watched_field: value}}`` from the workbook."""
    wb = _load(str(path), data_only=True)
    ws = wb.active
    if ws is None:
        return {}

    # Build header -> column index map (1-based, from openpyxl)
    headers: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value is None:
            continue
        headers[str(cell.value).strip()] = col_idx

    id_col = headers.get(_ID_COLUMN)
    if id_col is None:
        return {}

    fields_to_read = [(f, headers[f]) for f in _WATCHED_FIELDS if f in headers]

    out: dict[str, dict[str, Any]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        pid = ws.cell(row=r, column=id_col).value
        if not pid:
            continue
        pid = str(pid).strip()
        if not pid:
            continue
        out[pid] = {
            name: _normalize(ws.cell(row=r, column=c).value)
            for name, c in fields_to_read
        }
    return out


def save_snapshot(
    snapshot: dict[str, dict[str, Any]],
    cache_dir: Path | str = _SNAPSHOT_DIR,
    stamp: str | None = None,
) -> Path:
    """Persist ``snapshot`` as ``<cache_dir>/YYYY-MM-DD.json``. Returns the path."""
    cache_dir = Path(cache_dir)
    cache_dir.mkdir(parents=True, exist_ok=True)
    stamp = stamp or date.today().isoformat()
    target = cache_dir / f"{stamp}.json"
    payload = {
        "stamp": stamp,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "processes": snapshot,
    }
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return target


def load_latest_snapshot(
    cache_dir: Path | str = _SNAPSHOT_DIR,
    *,
    before: str | None = None,
) -> tuple[str, dict[str, dict[str, Any]]] | None:
    """Return the most recent saved snapshot (optionally strictly before ``before``).

    ``before`` is an ISO date string. Used on save to compare against the
    previous day's snapshot, not today's (if today's already exists).
    """
    cache_dir = Path(cache_dir)
    if not cache_dir.is_dir():
        return None
    candidates = sorted(cache_dir.glob("*.json"), reverse=True)
    for p in candidates:
        stamp = p.stem
        if before and stamp >= before:
            continue
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        processes = data.get("processes") or {}
        if isinstance(processes, dict):
            return stamp, processes
    return None


# ---------------------------------------------------------------------------
# Diffing
# ---------------------------------------------------------------------------
def diff_snapshots(
    old: dict[str, dict[str, Any]],
    new: dict[str, dict[str, Any]],
    *,
    prev_date: str = "",
) -> Changelog:
    """Compute added/removed/changed between two snapshots."""
    old_ids = set(old.keys())
    new_ids = set(new.keys())

    added = sorted(new_ids - old_ids)
    removed = sorted(old_ids - new_ids)
    changed: list[ProcessChange] = []

    for pid in sorted(old_ids & new_ids):
        old_row = old[pid] or {}
        new_row = new[pid] or {}
        field_changes: list[FieldChange] = []
        for f in _WATCHED_FIELDS:
            ov = _normalize(old_row.get(f))
            nv = _normalize(new_row.get(f))
            if ov != nv:
                # Ignore empty-to-empty transitions
                if ov in ("", None) and nv in ("", None):
                    continue
                field_changes.append(FieldChange(field=f, old=ov, new=nv))
        if field_changes:
            changed.append(ProcessChange(process_id=pid, changes=field_changes))

    return Changelog(prev_date=prev_date, added=added, removed=removed, changed=changed)


def summarize_changelog(cl: Changelog) -> str:
    """Produce a one-line human summary for the welcome card."""
    if cl.is_empty:
        return "Sin cambios desde la última corrida."

    parts: list[str] = []
    if cl.added:
        parts.append(f"{len(cl.added)} proceso(s) nuevo(s)")
    if cl.removed:
        parts.append(f"{len(cl.removed)} eliminado(s)")

    # Bucket the field-level changes by what moved
    state_changes = sum(
        1 for pc in cl.changed
        for c in pc.changes
        if c.field in ("Fase en SECOP", "Contrato: Estado")
    )
    mod_changes = sum(
        1 for pc in cl.changed
        for c in pc.changes
        if c.field in ("¿Hubo modificatorio?", "# modificatorios", "Docs: # Modificatorios")
    )
    money_changes = sum(
        1 for pc in cl.changed
        for c in pc.changes
        if c.field in ("Contrato: Valor", "Contrato: Valor pagado", "Proceso: Valor adjudicación", "Pagos: Total pagado")
    )

    if state_changes:
        parts.append(f"{state_changes} cambio(s) de estado")
    if mod_changes:
        parts.append(f"{mod_changes} cambio(s) en modificatorios")
    if money_changes:
        parts.append(f"{money_changes} cambio(s) de valor")
    return " · ".join(parts) if parts else "Sin cambios relevantes."


def _normalize(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


__all__ = [
    "Changelog",
    "FieldChange",
    "ProcessChange",
    "diff_snapshots",
    "load_latest_snapshot",
    "save_snapshot",
    "snapshot_from_excel",
    "summarize_changelog",
]
