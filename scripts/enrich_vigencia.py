"""Add ``vigencia`` and ``sheet`` to every item in
`.cache/watched_urls.json` based on the source Excel.

Used to backfill items imported BEFORE the vigencia field was added
to the importer. Reads the master Excel, builds a {url -> vigencia}
map across all 6 sheets, then patches the watch list in place.

Idempotent: items that already have vigencia are not overwritten
unless ``--force`` is given. Items whose URL is no longer in the
Excel keep their existing vigencia (and get ``sheet`` from their
``note`` if it has the "Importado de Excel · hoja XYZ" prefix).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from secop_ii.api import (
    _find_link_column,
    _find_vigencia_column,
    _vigencia_from_sheet_name,
)
from secop_ii.url_parser import InvalidSecopUrlError, parse_secop_url


def _build_excel_map(workbook_path: Path) -> dict[str, dict[str, str]]:
    """Return ``{process_id_or_url -> {vigencia, sheet}}`` from the Excel."""
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    out: dict[str, dict[str, str]] = {}
    sheet_summary: dict[str, int] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        link_loc = _find_link_column(ws)
        if not link_loc:
            sheet_summary[sheet_name] = 0
            continue
        header_row, link_col = link_loc
        vig_col = _find_vigencia_column(ws, header_row)
        sheet_fallback = _vigencia_from_sheet_name(sheet_name)

        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or len(row) < link_col:
                continue
            v = row[link_col - 1]
            if v is None:
                continue
            url = str(v).strip()
            if not url or "secop.gov.co" not in url.lower():
                continue

            vigencia: str | None = None
            if vig_col is not None and len(row) >= vig_col:
                rv = row[vig_col - 1]
                if rv is not None:
                    if isinstance(rv, (int, float)):
                        vigencia = str(int(rv))
                    else:
                        vigencia = str(rv).strip() or None
            if not vigencia:
                vigencia = sheet_fallback

            entry = {"vigencia": vigencia or "", "sheet": sheet_name}

            # Index by URL exact, by URL lower, and by parsed process_id
            out[url] = entry
            out[url.lower()] = entry
            try:
                ref = parse_secop_url(url)
                out[ref.process_id] = entry
            except InvalidSecopUrlError:
                pass
            count += 1
        sheet_summary[sheet_name] = count

    print(f"Excel map built. Hits per sheet: {sheet_summary}")
    print(f"Total keys in lookup: {len(out)}")
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook",
                   default="BASE DE DATOS FEAB CONTRATOS2.xlsx")
    ap.add_argument("--watch",
                   default=".cache/watched_urls.json")
    ap.add_argument("--force", action="store_true",
                   help="Overwrite vigencia on items that already have one.")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    watch_path = Path(args.watch)
    if not watch_path.exists():
        print(f"FATAL: {watch_path} not found")
        return 2

    items = json.loads(watch_path.read_text(encoding="utf-8"))
    print(f"Loaded {len(items)} items from {watch_path}")
    print()

    excel_map = _build_excel_map(Path(args.workbook))
    print()

    updated = 0
    no_match = 0
    for it in items:
        already = it.get("vigencia")
        if already and not args.force:
            continue
        # Try matching by URL, then URL.lower(), then process_id
        match = None
        url = it.get("url") or ""
        if url in excel_map:
            match = excel_map[url]
        elif url.lower() in excel_map:
            match = excel_map[url.lower()]
        elif it.get("process_id") and it["process_id"] in excel_map:
            match = excel_map[it["process_id"]]

        if match:
            it["vigencia"] = match["vigencia"] or None
            it["sheet"] = match["sheet"]
            updated += 1
        else:
            # Final fallback: parse vigencia from the existing note
            note = (it.get("note") or "")
            m = re.search(r"hoja\s+FEAB\s+(\d{4}(?:-\d{4})?)", note)
            if m:
                sheet_part = f"FEAB {m.group(1)}"
                it["sheet"] = sheet_part
                it["vigencia"] = (
                    _vigencia_from_sheet_name(sheet_part) or None
                )
                updated += 1
            else:
                no_match += 1

    if args.dry_run:
        print(f"DRY RUN: would update {updated}, no_match={no_match}")
        return 0

    watch_path.write_text(
        json.dumps(items, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Updated {updated} items with vigencia/sheet.")
    print(f"No match (kept as-is): {no_match}")

    # Distribution preview
    dist: dict[str, int] = {}
    for it in items:
        v = it.get("vigencia") or "(sin vigencia)"
        dist[v] = dist.get(v, 0) + 1
    print()
    print("Distribución por vigencia:")
    for k in sorted(dist.keys(), key=lambda s: (s == "(sin vigencia)", s),
                   reverse=False):
        print(f"  {k:>20s}: {dist[k]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
